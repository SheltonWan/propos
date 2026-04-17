#!/usr/bin/env python3
"""
PropOS 项目启动资料 - Excel 模板批量生成脚本

用法:
    python3 scripts/gen_kickoff_templates.py [输出目录]

默认输出到 docs/templates/kickoff/，每份资料单独一个文件。
"""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 颜色常量 ──────────────────────────────────────────────────────────
C_HEADER_REQUIRED   = "FF4A90D9"   # 必填列标题：蓝色
C_HEADER_OPTIONAL   = "FF9DC3E6"   # 选填列标题：浅蓝
C_SAMPLE            = "FFF2F2F2"   # 示例行背景：浅灰
C_INSTRUCTION       = "FFFFF2CC"   # 说明行背景：淡黄
C_BORDER            = "FFB8B8B8"   # 边框颜色

FONT_TITLE   = Font(name="微软雅黑", bold=True, size=13, color="FF1F3864")
FONT_HEADER  = Font(name="微软雅黑", bold=True, size=10, color="FFFFFFFF")
FONT_SAMPLE  = Font(name="微软雅黑", size=10, color="FF595959", italic=True)
FONT_BODY    = Font(name="微软雅黑", size=10)
FONT_NOTE    = Font(name="微软雅黑", size=9, color="FF7F7F7F", italic=True)

FILL_REQ   = PatternFill("solid", fgColor=C_HEADER_REQUIRED)
FILL_OPT   = PatternFill("solid", fgColor=C_HEADER_OPTIONAL)
FILL_SAMP  = PatternFill("solid", fgColor=C_SAMPLE)
FILL_INST  = PatternFill("solid", fgColor=C_INSTRUCTION)

thin_side = Side(border_style="thin", color=C_BORDER)
BORDER_ALL = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── 工具函数 ──────────────────────────────────────────────────────────

def new_wb() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def write_instruction(ws, row: int, text: str, col_span: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_NOTE
    cell.fill = FILL_INST
    cell.alignment = ALIGN_LEFT
    ws.row_dimensions[row].height = 22


def write_title(ws, row: int, text: str, col_span: int):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = FONT_TITLE
    cell.fill = PatternFill("solid", fgColor="FFD9E1F2")
    cell.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 32


def write_headers(ws, row: int, cols: list[tuple]):
    """cols: list of (header_text, col_width, is_required)"""
    for ci, (hdr, width, required) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=ci, value=hdr + (" ★" if required else ""))
        cell.font = FONT_HEADER
        cell.fill = FILL_REQ if required else FILL_OPT
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
        col_letter = get_column_letter(ci)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[row].height = 24


def write_sample_row(ws, row: int, values: list):
    for ci, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.font = FONT_SAMPLE
        cell.fill = FILL_SAMP
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER_ALL
    ws.row_dimensions[row].height = 20


def write_empty_rows(ws, start_row: int, count: int, col_count: int):
    for r in range(start_row, start_row + count):
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL
        ws.row_dimensions[r].height = 20


def freeze_header(ws, row: int, col: int = 1):
    ws.freeze_panes = ws.cell(row=row + 1, column=col)


def add_sheet(wb, title: str) -> object:
    ws = wb.create_sheet(title=title)
    return ws


# ── 01 写字楼单元台账 ────────────────────────────────────────────────

def build_office_units(wb):
    ws = add_sheet(wb, "写字楼单元台账")
    cols = [
        ("楼栋名称",     12, True),
        ("楼层编号",     10, True),
        ("单元编号",     14, True),
        ("建筑面积(m²)", 14, True),
        ("套内面积(m²)", 14, True),
        ("层高(m)",      10, False),
        ("朝向",         10, False),
        ("工位数",       10, False),
        ("分隔间数",     10, False),
        ("装修状态",     14, True),
        ("当前出租状态", 16, True),
        ("参考市场租金(元/m²/月)", 22, True),
        ("备注",         20, False),
    ]
    write_title(ws, 1, "PropOS 项目 — 写字楼单元台账导入模板", len(cols))
    write_instruction(ws, 2,
        "【填写说明】★ 为必填项。装修状态可填：毛坯/简装/精装/豪装。出租状态可填：空置/已租/装修中/非可租。多个分隔间用数字填写。",
        len(cols))
    write_headers(ws, 3, cols)
    samples = [
        ("A座", "5", "A-5F-01", 120.5, 115.0, 3.2, "南", 25, 3, "精装", "已租", 85, ""),
        ("A座", "5", "A-5F-02", 98.0, 94.0, 3.2, "北", 18, 2, "精装", "空置", 85, "待出租"),
        ("A座", "6", "A-6F-01", 200.0, 192.0, 3.2, "南北通透", 45, 5, "精装", "已租", 90, "大客户"),
    ]
    for i, s in enumerate(samples):
        write_sample_row(ws, 4 + i, list(s))
    write_empty_rows(ws, 4 + len(samples), 5, len(cols))
    freeze_header(ws, 3)
    return ws


# ── 02 商铺单元台账 ──────────────────────────────────────────────────

def build_retail_units(wb):
    ws = add_sheet(wb, "商铺单元台账")
    cols = [
        ("楼栋/区域",     14, True),
        ("楼层编号",      10, True),
        ("单元编号",      14, True),
        ("建筑面积(m²)",  14, True),
        ("套内面积(m²)",  14, True),
        ("门面宽度(m)",   12, False),
        ("是否临街",      10, False),
        ("层高(m)",       10, False),
        ("装修状态",      14, True),
        ("当前出租状态",  16, True),
        ("参考市场租金(元/m²/月)", 22, True),
        ("备注",          20, False),
    ]
    write_title(ws, 1, "PropOS 项目 — 商铺单元台账导入模板", len(cols))
    write_instruction(ws, 2,
        "【填写说明】★ 为必填项。是否临街填：是/否。装修状态可填：毛坯/简装/精装。出租状态可填：空置/已租/装修中/非可租。",
        len(cols))
    write_headers(ws, 3, cols)
    samples = [
        ("商铺区", "1", "S-1F-01", 85.0, 80.0, 6.5, "是", 4.5, "毛坯", "已租", 120, "沿街主铺"),
        ("商铺区", "1", "S-1F-02", 60.0, 56.0, 4.2, "否", 4.5, "简装", "空置", 100, ""),
        ("商铺区", "2", "S-2F-01", 120.0, 115.0, 5.0, "否", 4.0, "精装", "已租", 80, "餐饮类"),
    ]
    for i, s in enumerate(samples):
        write_sample_row(ws, 4 + i, list(s))
    write_empty_rows(ws, 4 + len(samples), 5, len(cols))
    freeze_header(ws, 3)


# ── 03 公寓单元台账 ──────────────────────────────────────────────────

def build_apartment_units(wb):
    ws = add_sheet(wb, "公寓单元台账")
    cols = [
        ("楼栋名称",     12, True),
        ("楼层编号",     10, True),
        ("单元编号",     14, True),
        ("建筑面积(m²)", 14, True),
        ("套内面积(m²)", 14, True),
        ("卧室数",       10, True),
        ("独立卫生间",   12, True),
        ("朝向",         10, False),
        ("装修状态",     14, True),
        ("当前出租状态", 16, True),
        ("参考市场租金(元/套/月)", 22, True),
        ("水电计费方式", 16, False),
        ("备注",         20, False),
    ]
    write_title(ws, 1, "PropOS 项目 — 公寓单元台账导入模板", len(cols))
    write_instruction(ws, 2,
        "【填写说明】★ 为必填项。独立卫生间填：是/否。水电计费方式填：含于租金/独立结算。出租状态可填：空置/已租/装修中/非可租。",
        len(cols))
    write_headers(ws, 3, cols)
    samples = [
        ("公寓楼", "3", "APT-3F-01", 65.0, 60.5, 1, "是", "南", "精装", "已租", 3500, "独立结算", ""),
        ("公寓楼", "3", "APT-3F-02", 88.0, 82.0, 2, "是", "南北", "精装", "空置", 4800, "独立结算", ""),
        ("公寓楼", "4", "APT-4F-01", 45.0, 42.0, 1, "否", "东", "简装", "已租", 2500, "含于租金", "一室一厅"),
    ]
    for i, s in enumerate(samples):
        write_sample_row(ws, 4 + i, list(s))
    write_empty_rows(ws, 4 + len(samples), 5, len(cols))
    freeze_header(ws, 3)


# ── 04 楼栋基本信息 ──────────────────────────────────────────────────

def build_buildings(wb):
    ws = add_sheet(wb, "楼栋基本信息")
    cols = [
        ("楼栋名称",       16, True),
        ("业态类型",       14, True),
        ("总楼层数",       12, True),
        ("地上楼层",       12, False),
        ("地下楼层",       12, False),
        ("建筑面积GFA(m²)",20, True),
        ("可租面积NLA(m²)",20, True),
        ("竣工年份",       12, False),
        ("物业地址",       30, False),
        ("备注",           20, False),
    ]
    write_title(ws, 1, "PropOS 项目 — 楼栋基本信息", len(cols))
    write_instruction(ws, 2,
        "【填写说明】业态类型可填：写字楼/商铺/公寓/混合。每个楼栋填写一行。",
        len(cols))
    write_headers(ws, 3, cols)
    samples = [
        ("A座", "写字楼", 20, 20, 2, 30000, 27000, 2015, "XX市XX路XX号A座", "主楼"),
        ("商铺区", "商铺", 3, 3, 1, 2707, 2500, 2015, "XX市XX路XX号商铺区", ""),
        ("公寓楼", "公寓", 15, 15, 1, 7293, 6800, 2016, "XX市XX路XX号公寓楼", ""),
    ]
    for i, s in enumerate(samples):
        write_sample_row(ws, 4 + i, list(s))
    write_empty_rows(ws, 4 + len(samples), 3, len(cols))
    freeze_header(ws, 3)


# ── 05 图纸版本说明 & 单元编号对照表 ────────────────────────────────

def build_drawings(wb):
    # Sheet1: 图纸版本说明
    ws1 = add_sheet(wb, "图纸版本说明")
    cols1 = [
        ("楼栋名称",     14, True),
        ("楼层编号",     12, True),
        ("图纸文件名",   30, True),
        ("版本日期",     14, True),
        ("是否当前生效", 14, True),
        ("修改原因说明", 30, False),
        ("上传人",       14, False),
        ("备注",         20, False),
    ]
    write_title(ws1, 1, "PropOS 项目 — 图纸版本说明", len(cols1))
    write_instruction(ws1, 2,
        "【填写说明】每个楼层的每个图纸版本填一行。是否当前生效填：是/否。版本日期格式：YYYY-MM-DD。",
        len(cols1))
    write_headers(ws1, 3, cols1)
    samples1 = [
        ("A座", "5", "A座5层平面图_v2023.dwg", "2023-06-15", "是", "2023年改造后更新", "张工", ""),
        ("A座", "5", "A座5层平面图_v2021.dwg", "2021-03-01", "否", "原始版本", "李工", "已归档"),
        ("A座", "6", "A座6层平面图_v2022.dwg", "2022-09-20", "是", "首次提供", "张工", ""),
    ]
    for i, s in enumerate(samples1):
        write_sample_row(ws1, 4 + i, list(s))
    write_empty_rows(ws1, 4 + len(samples1), 5, len(cols1))
    freeze_header(ws1, 3)

    # Sheet2: 单元编号对照表
    ws2 = add_sheet(wb, "单元编号对照表")
    cols2 = [
        ("楼栋名称",        14, True),
        ("楼层编号",        12, True),
        ("CAD图纸中标注编号", 22, True),
        ("台账中的单元编号",  22, True),
        ("备注",            20, False),
    ]
    write_title(ws2, 1, "PropOS 项目 — CAD图纸与台账单元编号对照表", len(cols2))
    write_instruction(ws2, 2,
        "【填写说明】如图纸中编号与台账编号完全相同可不填此表。若存在差异（如图纸标'501'，台账标'A-5F-01'），必须逐条填写映射关系。",
        len(cols2))
    write_headers(ws2, 3, cols2)
    samples2 = [
        ("A座", "5", "501", "A-5F-01", ""),
        ("A座", "5", "502", "A-5F-02", ""),
        ("A座", "6", "601A", "A-6F-01", "原601已拆分为601A和601B"),
    ]
    for i, s in enumerate(samples2):
        write_sample_row(ws2, 4 + i, list(s))
    write_empty_rows(ws2, 4 + len(samples2), 5, len(cols2))
    freeze_header(ws2, 3)


# ── 06 合同清单 & 租金递增规则 ──────────────────────────────────────

def build_contracts(wb):
    # Sheet1: 有效合同清单
    ws1 = add_sheet(wb, "当前有效合同清单")
    cols1 = [
        ("合同编号",         16, True),
        ("租客名称",         20, True),
        ("业态类型",         12, True),
        ("关联单元编号(多个用|分隔)", 30, True),
        ("起租日",           14, True),
        ("到期日",           14, True),
        ("计费面积(m²)",     16, True),
        ("月租金单价(元/m²)", 18, True),
        ("月租金总额(元)",   16, True),
        ("含税/不含税",      14, True),
        ("适用税率(%)",      14, True),
        ("押金金额(元)",     16, True),
        ("付款周期",         12, True),
        ("免租期开始日",     16, False),
        ("免租期结束日",     16, False),
        ("合同状态",         14, True),
        ("备注",             20, False),
    ]
    write_title(ws1, 1, "PropOS 项目 — 当前有效合同清单", len(cols1))
    write_instruction(ws1, 2,
        "【填写说明】付款周期填：月付/季付/半年付/年付。合同状态填：执行中/即将到期（<=90天）。含税/不含税填：含税/不含税。",
        len(cols1))
    write_headers(ws1, 3, cols1)
    samples1 = [
        ("HT-2024-001", "XX科技有限公司", "写字楼", "A-5F-01|A-5F-02", "2024-03-01", "2027-02-28",
         213.5, 85, 18147.5, "含税", 9, 54442, "季付", "", "", "执行中", ""),
        ("HT-2024-002", "YY传媒集团", "写字楼", "A-6F-01", "2024-06-01", "2026-05-31",
         200.0, 90, 18000, "不含税", 9, 54000, "月付", "2024-06-01", "2024-06-14", "即将到期", "到期前90天"),
        ("HT-2025-001", "ZZ餐饮有限公司", "商铺", "S-1F-01", "2025-01-01", "2030-12-31",
         80.0, 120, 9600, "含税", 9, 28800, "季付", "2025-01-01", "2025-03-31", "执行中", "有分成条款"),
    ]
    for i, s in enumerate(samples1):
        write_sample_row(ws1, 4 + i, list(s))
    write_empty_rows(ws1, 4 + len(samples1), 5, len(cols1))
    freeze_header(ws1, 3)

    # Sheet2: 租金递增规则
    ws2 = add_sheet(wb, "租金递增规则")
    cols2 = [
        ("合同编号",             16, True),
        ("阶段序号",             10, True),
        ("阶段开始日期",         16, True),
        ("阶段结束日期",         16, True),
        ("递增类型",             20, True),
        ("固定递增比例(%)",      18, False),
        ("固定递增金额(元/m²)",  20, False),
        ("递增间隔(年)",         14, False),
        ("阶梯租金(若阶梯类型)", 30, False),
        ("CPI年份(若CPI类型)",   18, False),
        ("备注",                 24, False),
    ]
    write_title(ws2, 1, "PropOS 项目 — 租金递增规则配置", len(cols2))
    write_instruction(ws2, 2,
        "【填写说明】递增类型填：固定比例递增/固定金额递增/阶梯式递增/CPI挂钩递增/每N年递增/免租后基准调整。阶梯租金格式：第1年80|第2年85|第3年90。",
        len(cols2))
    write_headers(ws2, 3, cols2)
    samples2 = [
        ("HT-2024-001", 1, "2024-03-01", "2026-02-28", "固定比例递增", 5, "", 1, "", "", "每年涨5%"),
        ("HT-2024-001", 2, "2026-03-01", "2027-02-28", "CPI挂钩递增", "", "", 1, "", 2025, "按上年CPI调整"),
        ("HT-2025-001", 1, "2025-04-01", "2030-12-31", "阶梯式递增", "", "", "", "第1-2年120|第3-4年130|第5年140", "", "商铺阶梯"),
    ]
    for i, s in enumerate(samples2):
        write_sample_row(ws2, 4 + i, list(s))
    write_empty_rows(ws2, 4 + len(samples2), 5, len(cols2))
    freeze_header(ws2, 3)


# ── 07 未结账单 ──────────────────────────────────────────────────────

def build_invoices(wb):
    ws = add_sheet(wb, "当前未结账单")
    cols = [
        ("账单编号",         16, True),
        ("合同编号",         16, True),
        ("租客名称",         20, True),
        ("费项类型",         16, True),
        ("账期年月",         12, True),
        ("账期截止日",       14, True),
        ("应收金额含税(元)", 18, True),
        ("应收金额不含税(元)", 20, True),
        ("适用税率(%)",      14, True),
        ("账单状态",         14, True),
        ("已收金额(元)",     16, False),
        ("最近收款日期",     16, False),
        ("发票状态",         14, False),
        ("发票号码",         18, False),
        ("备注",             20, False),
    ]
    write_title(ws, 1, "PropOS 项目 — 当前未结账单导入模板", len(cols))
    write_instruction(ws, 2,
        "【填写说明】费项类型填：租金/物管费/水费/电费/停车费/储藏室费/其他。账单状态填：未收/部分收款/已收。发票状态填：未开/已开。",
        len(cols))
    write_headers(ws, 3, cols)
    samples = [
        ("INV-2026-0401", "HT-2024-001", "XX科技有限公司", "租金", "2026-04", "2026-04-30",
         18147.5, 16649.08, 9, "未收", 0, "", "未开", "", ""),
        ("INV-2026-0402", "HT-2024-001", "XX科技有限公司", "物管费", "2026-04", "2026-04-30",
         2135, 1959.63, 9, "未收", 0, "", "未开", "", ""),
        ("INV-2026-0403", "HT-2024-002", "YY传媒集团", "租金", "2026-04", "2026-04-30",
         18000, 16513.76, 9, "部分收款", 10000, "2026-04-10", "已开", "FP20260001", "欠款8000"),
        ("INV-2026-0301", "HT-2024-002", "YY传媒集团", "租金", "2026-03", "2026-03-31",
         18000, 16513.76, 9, "部分收款", 5000, "2026-03-20", "已开", "FP20260002", "逾期账单"),
    ]
    for i, s in enumerate(samples):
        write_sample_row(ws, 4 + i, list(s))
    write_empty_rows(ws, 4 + len(samples), 5, len(cols))
    freeze_header(ws, 3)


# ── 08 押金台账 & 收款核销记录 ──────────────────────────────────────

def build_deposits_payments(wb):
    # Sheet1: 押金台账
    ws1 = add_sheet(wb, "押金台账")
    cols1 = [
        ("合同编号",       16, True),
        ("租客名称",       20, True),
        ("押金金额(元)",   16, True),
        ("收取日期",       14, True),
        ("当前状态",       14, True),
        ("冻结金额(元)",   16, False),
        ("冲抵金额(元)",   16, False),
        ("冲抵原因",       30, False),
        ("退还金额(元)",   16, False),
        ("退还日期",       14, False),
        ("退还方式",       14, False),
        ("备注",           20, False),
    ]
    write_title(ws1, 1, "PropOS 项目 — 押金台账", len(cols1))
    write_instruction(ws1, 2,
        "【填写说明】当前状态填：已收取/冻结中/部分冲抵/已退还。退还方式填：原路退回/银行转账。每份合同的押金填一行。",
        len(cols1))
    write_headers(ws1, 3, cols1)
    samples1 = [
        ("HT-2024-001", "XX科技有限公司", 54442, "2024-02-20", "已收取", 0, 0, "", 0, "", "", ""),
        ("HT-2024-002", "YY传媒集团", 54000, "2024-05-15", "部分冲抵", 0, 5000, "欠缴3月租金抵扣", 0, "", "", ""),
        ("HT-2023-005", "WW贸易公司", 36000, "2023-01-01", "已退还", 0, 2000, "维修损坏", 34000, "2026-01-15", "银行转账", "已终止合同"),
    ]
    for i, s in enumerate(samples1):
        write_sample_row(ws1, 4 + i, list(s))
    write_empty_rows(ws1, 4 + len(samples1), 5, len(cols1))
    freeze_header(ws1, 3)

    # Sheet2: 收款核销记录
    ws2 = add_sheet(wb, "收款核销记录")
    cols2 = [
        ("核销单号",         16, True),
        ("关联账单编号(多个用|分隔)", 30, True),
        ("租客名称",         20, True),
        ("收款日期",         14, True),
        ("收款金额(元)",     16, True),
        ("收款账户/银行",    20, False),
        ("银行流水号",       24, False),
        ("核销方式",         14, True),
        ("核销说明",         30, False),
        ("经办人",           14, False),
    ]
    write_title(ws2, 1, "PropOS 项目 — 收款核销记录", len(cols2))
    write_instruction(ws2, 2,
        "【填写说明】核销方式填：全额核销/部分核销/跨账单核销。若一次收款核销多张账单，账单编号用'|'分隔。",
        len(cols2))
    write_headers(ws2, 3, cols2)
    samples2 = [
        ("CX-2026-001", "INV-2026-0401|INV-2026-0402", "XX科技有限公司", "2026-04-05", 20282.5, "建设银行", "CX20260001", "全额核销", "季度款一次收清", "李财务"),
        ("CX-2026-002", "INV-2026-0403", "YY传媒集团", "2026-04-10", 10000, "工商银行", "CX20260002", "部分核销", "先收一半，月底补足", "李财务"),
    ]
    for i, s in enumerate(samples2):
        write_sample_row(ws2, 4 + i, list(s))
    write_empty_rows(ws2, 4 + len(samples2), 5, len(cols2))
    freeze_header(ws2, 3)


# ── 09 租客主档 ──────────────────────────────────────────────────────

def build_tenants(wb):
    # Sheet1: 企业租客
    ws1 = add_sheet(wb, "企业租客清单")
    cols1 = [
        ("企业名称",               24, True),
        ("统一社会信用代码",       24, True),
        ("联系人姓名",             16, True),
        ("联系人手机",             16, True),
        ("联系人邮箱",             24, False),
        ("紧急联系人",             16, False),
        ("紧急联系电话",           16, False),
        ("当前合同编号(多个用|分隔)", 30, True),
        ("开票信息-抬头",          24, False),
        ("开票信息-纳税人识别号",  24, False),
        ("开票信息-开户行",        24, False),
        ("开票信息-银行账号",      24, False),
        ("备注",                   20, False),
    ]
    write_title(ws1, 1, "PropOS 项目 — 企业租客清单", len(cols1))
    write_instruction(ws1, 2,
        "【填写说明】★ 为必填项。统一社会信用代码用于唯一标识企业。开票信息可在此提供，以便开具增值税发票。",
        len(cols1))
    write_headers(ws1, 3, cols1)
    samples1 = [
        ("XX科技有限公司", "91310000XXXXXXXXXX", "王小明", "138XXXXXXXX", "wangxm@xxtech.com",
         "张总", "139XXXXXXXX", "HT-2024-001", "XX科技有限公司", "91310000XXXXXXXXXX", "中国建设银行XX支行", "1234567890123456", ""),
        ("YY传媒集团有限公司", "91310000YYYYYYYYYY", "李小红", "136XXXXXXXX", "lixh@yymedia.com",
         "赵总", "137XXXXXXXX", "HT-2024-002", "", "", "", "", "逾期风险，需关注"),
    ]
    for i, s in enumerate(samples1):
        write_sample_row(ws1, 4 + i, list(s))
    write_empty_rows(ws1, 4 + len(samples1), 5, len(cols1))
    freeze_header(ws1, 3)

    # Sheet2: 个人租客（公寓）
    ws2 = add_sheet(wb, "个人租客清单（公寓）")
    cols2 = [
        ("姓名",               14, True),
        ("身份证号(后4位仅示例)", 28, True),
        ("联系手机",           16, True),
        ("联系邮箱",           24, False),
        ("紧急联系人",         16, False),
        ("紧急联系电话",       16, False),
        ("当前合同编号",       18, True),
        ("入住人数",           10, False),
        ("备注",               20, False),
    ]
    write_title(ws2, 1, "PropOS 项目 — 个人租客清单（公寓）", len(cols2))
    write_instruction(ws2, 2,
        "【重要：个人信息保护】身份证号为敏感信息，系统将加密存储，API层默认脱敏（仅显示后4位）。提供前请确认已在租赁合同中获得数据处理授权。",
        len(cols2))
    write_headers(ws2, 3, cols2)
    samples2 = [
        ("张三", "310XXX########1234", "135XXXXXXXX", "", "张父", "136XXXXXXXX", "HT-APT-2025-001", 2, ""),
        ("李四", "440XXX########5678", "139XXXXXXXX", "li4@email.com", "", "", "HT-APT-2025-002", 1, "单人入住"),
    ]
    for i, s in enumerate(samples2):
        write_sample_row(ws2, 4 + i, list(s))
    write_empty_rows(ws2, 4 + len(samples2), 5, len(cols2))
    freeze_header(ws2, 3)


# ── 10 组织架构 & 员工账号 & 管辖范围 ─────────────────────────────

def build_org(wb):
    # Sheet1: 组织架构树
    ws1 = add_sheet(wb, "组织架构树")
    cols1 = [
        ("层级",           10, True),
        ("节点名称",       24, True),
        ("上级节点名称",   24, True),
        ("负责人姓名",     16, False),
        ("备注",           20, False),
    ]
    write_title(ws1, 1, "PropOS 项目 — 组织架构树（最多3级）", len(cols1))
    write_instruction(ws1, 2,
        "【填写说明】层级填：1（公司）/2（部门）/3（组）。顶层节点的上级节点名称留空。最多支持3级组织树。",
        len(cols1))
    write_headers(ws1, 3, cols1)
    samples1 = [
        (1, "XX物业管理有限公司", "", "王总", "顶层"),
        (2, "租务部", "XX物业管理有限公司", "陈经理", ""),
        (2, "财务部", "XX物业管理有限公司", "李财务总监", ""),
        (2, "物业运营部", "XX物业管理有限公司", "张经理", ""),
        (3, "租务一组", "租务部", "小王", "负责写字楼"),
        (3, "租务二组", "租务部", "小李", "负责公寓+商铺"),
    ]
    for i, s in enumerate(samples1):
        write_sample_row(ws1, 4 + i, list(s))
    write_empty_rows(ws1, 4 + len(samples1), 3, len(cols1))
    freeze_header(ws1, 3)

    # Sheet2: 员工账号申请
    ws2 = add_sheet(wb, "员工账号申请")
    role_note = "角色说明：super_admin=超级管理员 | operations_manager=运营管理层 | leasing_specialist=租务专员 | finance_staff=财务人员 | maintenance_staff=维修技工 | property_inspector=楼管巡检 | report_viewer=只读观察者"
    cols2 = [
        ("姓名",         14, True),
        ("所属部门",     20, True),
        ("岗位角色",     28, True),
        ("登录邮箱",     28, True),
        ("联系手机",     16, True),
        ("是否管理员",   12, False),
        ("备注",         20, False),
    ]
    write_title(ws2, 1, "PropOS 项目 — 员工账号申请表", len(cols2))
    write_instruction(ws2, 2, f"【角色说明】{role_note}", len(cols2))
    write_headers(ws2, 3, cols2)
    samples2 = [
        ("王总", "XX物业管理有限公司", "super_admin", "wangzong@company.com", "138XXXXXXXX", "是", "股东兼CEO"),
        ("陈经理", "租务部", "operations_manager", "chenjl@company.com", "136XXXXXXXX", "否", ""),
        ("小王", "租务部", "leasing_specialist", "xiaoWang@company.com", "135XXXXXXXX", "否", ""),
        ("李财务", "财务部", "finance_staff", "licw@company.com", "137XXXXXXXX", "否", ""),
        ("维修张", "物业运营部", "maintenance_staff", "zhangwx@company.com", "139XXXXXXXX", "否", ""),
        ("巡检刘", "物业运营部", "property_inspector", "liuxj@company.com", "132XXXXXXXX", "否", ""),
        ("投资人赵", "", "report_viewer", "zhao@investor.com", "150XXXXXXXX", "否", "只读观察者"),
    ]
    for i, s in enumerate(samples2):
        write_sample_row(ws2, 4 + i, list(s))
    write_empty_rows(ws2, 4 + len(samples2), 5, len(cols2))
    freeze_header(ws2, 3)

    # Sheet3: 管辖范围配置
    ws3 = add_sheet(wb, "管辖范围配置")
    cols3 = [
        ("员工姓名/部门名称", 22, True),
        ("类型(员工/部门)",    16, True),
        ("管辖楼栋",          20, True),
        ("管辖楼层范围",      20, False),
        ("管辖业态",          20, False),
        ("备注",              30, False),
    ]
    write_title(ws3, 1, "PropOS 项目 — KPI管辖范围配置", len(cols3))
    write_instruction(ws3, 2,
        "【填写说明】管辖楼栋填楼栋名称，多个用|分隔。管辖楼层填如'5-10层'或'全部'。管辖业态填：写字楼/商铺/公寓/全部。个人配置优先于部门配置。",
        len(cols3))
    write_headers(ws3, 3, cols3)
    samples3 = [
        ("租务部", "部门", "A座|商铺区", "全部", "写字楼|商铺", "部门默认范围"),
        ("物业运营部", "部门", "A座|商铺区|公寓楼", "全部", "全部", ""),
        ("小王", "员工", "A座", "5-10层", "写字楼", "个人覆盖（仅负责A座5-10层）"),
    ]
    for i, s in enumerate(samples3):
        write_sample_row(ws3, 4 + i, list(s))
    write_empty_rows(ws3, 4 + len(samples3), 5, len(cols3))
    freeze_header(ws3, 3)


# ── 11 二房东账号申请 ────────────────────────────────────────────────

def build_sublord(wb):
    ws = add_sheet(wb, "二房东账号申请")
    cols = [
        ("二房东企业名称",     24, True),
        ("主联系人姓名",       16, True),
        ("主联系人手机",       16, True),
        ("主联系人邮箱",       28, True),
        ("绑定的主合同编号",   24, True),
        ("覆盖楼栋名称",       20, True),
        ("覆盖楼层范围",       20, True),
        ("账号有效期",         16, True),
        ("备注",               20, False),
    ]
    write_title(ws, 1, "PropOS 项目 — 二房东账号申请表", len(cols))
    write_instruction(ws, 2,
        "【填写说明】账号有效期通常与主合同到期日一致。一个账号可绑定多份主合同，合同编号用|分隔。系统会在主合同到期后自动冻结账号。",
        len(cols))
    write_headers(ws, 3, cols)
    samples = [
        ("二房东A公司", "张负责人", "138XXXXXXXX", "zhangfzr@sublord-a.com", "HT-2024-001|HT-2024-003", "A座", "5-10层", "2027-02-28", "大面积转租方"),
        ("二房东B公司", "李负责人", "136XXXXXXXX", "lifzr@sublord-b.com", "HT-2025-010", "公寓楼", "3-8层", "2027-12-31", "公寓整层转租"),
    ]
    for i, s in enumerate(samples):
        write_sample_row(ws, 4 + i, list(s))
    write_empty_rows(ws, 4 + len(samples), 3, len(cols))
    freeze_header(ws, 3)


# ── 12 KPI 方案设计 ──────────────────────────────────────────────────

def build_kpi(wb):
    # Sheet1: KPI方案设计
    ws1 = add_sheet(wb, "KPI方案设计")
    cols1 = [
        ("方案名称",           24, True),
        ("适用部门或员工",     24, True),
        ("评估周期",           14, True),
        ("方案生效日期",       16, True),
        ("方案终止日期",       16, False),
        ("方案说明",           30, False),
    ]
    write_title(ws1, 1, "PropOS 项目 — KPI方案设计（方案基本信息）", len(cols1))
    write_instruction(ws1, 2,
        "【填写说明】评估周期填：月度/季度/年度。一个方案填一行。指标权重在下一个Sheet中填写。",
        len(cols1))
    write_headers(ws1, 3, cols1)
    samples1 = [
        ("租务部2026考核方案", "租务部", "季度", "2026-01-01", "2026-12-31", "适用租务部全员"),
        ("物业运营部2026方案", "物业运营部", "月度", "2026-01-01", "2026-12-31", "含工单响应指标"),
        ("陈经理个人方案2026", "陈经理", "年度", "2026-01-01", "2026-12-31", "管理层个人方案"),
    ]
    for i, s in enumerate(samples1):
        write_sample_row(ws1, 4 + i, list(s))
    write_empty_rows(ws1, 4 + len(samples1), 3, len(cols1))
    freeze_header(ws1, 3)

    # Sheet2: KPI指标权重配置
    ws2 = add_sheet(wb, "KPI指标权重配置")
    cols2 = [
        ("所属方案名称",               24, True),
        ("指标编号",                   10, True),
        ("指标名称",                   24, True),
        ("是否启用",                   10, True),
        ("权重(%，同方案合计=100)",    24, True),
        ("满分阈值(默认见说明)",       20, False),
        ("及格阈值",                   14, False),
        ("备注",                       20, False),
    ]
    write_title(ws2, 1, "PropOS 项目 — KPI指标权重配置", len(cols2))
    write_instruction(ws2, 2,
        "【填写说明】同一方案内所有启用指标的权重之和必须=100%。满分阈值不填则用系统默认值。是否启用填：是/否。",
        len(cols2))
    write_headers(ws2, 3, cols2)
    samples2 = [
        ("租务部2026考核方案", "K01", "出租率", "是", 20, "≥95%", "≥85%", ""),
        ("租务部2026考核方案", "K02", "收款及时率", "是", 15, "≥95%", "≥80%", ""),
        ("租务部2026考核方案", "K04", "续约率", "是", 20, "≥80%", "≥60%", ""),
        ("租务部2026考核方案", "K06", "空置周转天数", "是", 15, "≤30天", "≤60天", "反向指标"),
        ("租务部2026考核方案", "K08", "逾期率", "是", 15, "≤5%", "≤15%", "反向指标"),
        ("租务部2026考核方案", "K13", "新签约面积", "是", 15, "≥2000m²", "≥1000m²", ""),
        ("租务部2026考核方案", "K10", "租户满意度", "是", "注：以上合计100%", "≥90分", "≥75分", "手动录入"),
    ]
    for i, s in enumerate(samples2):
        write_sample_row(ws2, 4 + i, list(s))
    write_empty_rows(ws2, 4 + len(samples2), 3, len(cols2))
    freeze_header(ws2, 3)

    # Sheet3: KPI指标参考说明
    ws3 = add_sheet(wb, "KPI指标说明（参考）")
    cols3 = [
        ("指标编号", 10, False),
        ("指标名称", 24, False),
        ("指标方向", 12, False),
        ("数据来源", 20, False),
        ("默认满分标准", 20, False),
        ("计算说明", 50, False),
    ]
    write_title(ws3, 1, "系统预定义 KPI 指标库（参考，不需客户填写）", len(cols3))
    write_instruction(ws3, 2,
        "【说明】本表为系统预定义指标参考，客户只需在'KPI指标权重配置'表中配置即可，无需修改本表。",
        len(cols3))
    write_headers(ws3, 3, cols3)
    ref_data = [
        ("K01", "出租率", "正向", "资产台账", "≥95%", "已租NLA ÷ 总可租NLA"),
        ("K02", "收款及时率", "正向", "账单核销", "≥95%", "账期内到账笔数 ÷ 应收笔数"),
        ("K03", "租户集中度", "反向（越低越好）", "合同数据", "≤40%", "前3大租户租金占比"),
        ("K04", "续约率", "正向", "合同状态机", "≥80%", "续签合同数 ÷ 到期合同数"),
        ("K05", "工单响应时效", "反向（越低越好）", "工单系统", "≤24小时", "仅repair类型：平均派单到完工时长"),
        ("K06", "空置周转天数", "反向（越低越好）", "资产+合同", "≤30天", "单元空置至重新签约的平均天数"),
        ("K07", "NOI达成率", "正向", "NOI看板", "≥100%", "实际NOI ÷ 预算NOI"),
        ("K08", "逾期率", "反向（越低越好）", "账单数据", "≤5%", "逾期金额 ÷ 应收金额"),
        ("K09", "租金递增执行率", "正向", "递增配置器", "≥95%", "按递增规则成功调价合同数 ÷ 应调价合同数"),
        ("K10", "租户满意度", "正向", "手动录入", "≥90分", "管理员按季度手动录入调查评分"),
        ("K11", "预防性维修率", "正向", "工单系统", "≥90%", "预防性维修工单数 ÷ 总工单数"),
        ("K12", "空置面积降幅", "正向", "资产台账", "≥20%", "（上期空置 - 本期空置）÷ 上期空置"),
        ("K13", "新签约面积", "正向", "合同数据", "≥2000m²", "本期新签合同关联单元总面积"),
        ("K14", "续签率", "正向", "合同状态机", "≥80%", "到期90天内续签合同数 ÷ 到期合同数"),
    ]
    for i, row in enumerate(ref_data):
        cell_row = 4 + i
        for ci, val in enumerate(row, 1):
            cell = ws3.cell(row=cell_row, column=ci, value=val)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL
        ws3.row_dimensions[cell_row].height = 20
    freeze_header(ws3, 3)


# ── 13 业务规则口径确认表 ─────────────────────────────────────────

def build_rules_confirm(wb):
    ws = add_sheet(wb, "业务规则口径确认")
    cols = [
        ("规则类别",            16, False),
        ("确认问题",            40, False),
        ("系统默认值",          30, False),
        ("客户确认选择",        30, True),
        ("客户补充说明",        30, False),
        ("确认人",              14, True),
        ("确认日期",            14, True),
    ]
    write_title(ws, 1, "PropOS 项目 — 业务规则口径确认表（必须返回）", len(cols))
    write_instruction(ws, 2,
        "【重要】请在<客户确认选择>列填写您的选择，<确认人>和<确认日期>为必填。此表返回后将作为系统配置依据，后续变更需走正式变更流程。",
        len(cols))
    write_headers(ws, 3, cols)
    rules = [
        ("合同计费", "首期/末期不足月，是否按自然日折算？", "默认按自然日折算（月标准金额÷当月自然日×实际天数）", "", "", "", ""),
        ("合同计费", "写字楼/商铺按建面还是套内计费？", "默认按合同约定计费面积（需在合同中明确标注）", "", "", "", ""),
        ("税费口径", "各类租户适用增值税税率？", "一般纳税人9%/简易征收5%，按合同单独标注", "", "", "", ""),
        ("押金规则", "押金是否允许转移至续签合同（无需先退后收）？", "允许转移", "", "", "", ""),
        ("押金规则", "押金冲抵违约金/欠费前，是否需要线上审批？", "需要财务人员确认无欠费后审批", "", "", "", ""),
        ("收款核销", "多账单时，核销优先顺序？", "默认先到期先核销，允许财务手工调整", "", "", "", ""),
        ("逾期判断", "账期截止后第几天开始计算逾期？", "截止日后第1天即计逾期", "", "", "", ""),
        ("预警节点", "到期预警：提前90/60/30天是否全部触发？", "三个节点全部发送", "", "", "", ""),
        ("预警接收", "合同到期预警的具体接收邮箱/人员？", "租务专员+运营管理层（根据系统账号自动匹配）", "", "", "", ""),
        ("预警接收", "逾期账单提醒的具体接收邮箱/人员？", "财务人员+租务专员", "", "", "", ""),
        ("KPI申诉", "快照冻结后，允许员工申诉的天数？", "7个自然日", "", "", "", ""),
        ("NOI口径", "水电代收部分，是否计入EGI收入？", "以客户实际财务处理方式为准", "", "", "", ""),
        ("NOI口径", "外包物业管理费，是否作为代收科目？", "以客户实际财务处理方式为准", "", "", "", ""),
        ("CapEx阈值", "单次维修超过多少元才列为资本性支出(CapEx，不进NOI)？", "默认单次>5,000元且延长资产使用寿命", "", "", "", ""),
        ("水电计价", "各业态电价（元/度）是多少？是否有阶梯电价？", "客户提供具体单价", "", "", "", ""),
        ("水电计价", "各业态水价（元/吨）是多少？是否有阶梯水价？", "客户提供具体单价", "", "", "", ""),
        ("商铺分成", "保底租金+营业额分成，各商铺分成比例是否相同？", "每份合同单独配置", "", "", "", ""),
        ("二房东审核", "二房东提交子租赁数据后，内部审核SLA（工作日）？", "2个工作日", "", "", "", ""),
        ("二房东提报", "二房东每月数据提报截止日期？", "每月5日前完成上月数据填报", "", "", "", ""),
    ]
    for i, row in enumerate(rules):
        cell_row = 4 + i
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=cell_row, column=ci, value=val)
            if ci == 4:  # 客户确认列高亮
                cell.fill = PatternFill("solid", fgColor="FFFFF2CC")
                cell.font = FONT_BODY
            elif ci in (6, 7):  # 确认人/日期
                cell.fill = PatternFill("solid", fgColor="FFFFE2CC")
                cell.font = FONT_BODY
            else:
                cell.font = FONT_BODY
            cell.alignment = ALIGN_LEFT
            cell.border = BORDER_ALL
        ws.row_dimensions[cell_row].height = 30
    freeze_header(ws, 3)


# ── 主流程 ─────────────────────────────────────────────────────────

TEMPLATE_DEFS = [
    ("01_资产台账_写字楼商铺公寓", lambda wb: [build_office_units(wb), build_retail_units(wb), build_apartment_units(wb), build_buildings(wb)]),
    ("02_图纸版本说明_单元对照表", lambda wb: build_drawings(wb)),
    ("03_合同清单_租金递增规则",   lambda wb: build_contracts(wb)),
    ("04_账单_押金_收款核销",      lambda wb: [build_invoices(wb), build_deposits_payments(wb)]),
    ("05_租客主档_企业_个人",      lambda wb: build_tenants(wb)),
    ("06_组织架构_员工账号_管辖范围", lambda wb: build_org(wb)),
    ("07_二房东账号申请",          lambda wb: build_sublord(wb)),
    ("08_KPI方案设计_指标权重",    lambda wb: build_kpi(wb)),
    ("09_业务规则口径确认（必须返回）", lambda wb: build_rules_confirm(wb)),
]


def main():
    out_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("docs/templates/kickoff")
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"\nPropOS 项目启动资料模板生成器")
    print(f"输出目录: {out_dir.resolve()}\n")

    for filename, builder in TEMPLATE_DEFS:
        wb = new_wb()
        builder(wb)
        out_path = out_dir / f"{filename}.xlsx"
        wb.save(str(out_path))
        print(f"  ✅  {filename}.xlsx")

    print(f"\n共生成 {len(TEMPLATE_DEFS)} 个模板文件。")
    print("填写时请注意：★ 标注的列为必填项；黄色底色的单元格需要客户填写确认内容。\n")


if __name__ == "__main__":
    main()
