"""生成二房东子租赁数据导入模板 (10_子租赁数据导入_租客与合同.xlsx)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL   = PatternFill("solid", fgColor="2F5EB3")
REQUIRED_FILL = PatternFill("solid", fgColor="FFF2CC")
OPTIONAL_FILL = PatternFill("solid", fgColor="EBF5EB")
EXAMPLE_FILL  = PatternFill("solid", fgColor="F2F2F2")
INFO_FILL     = PatternFill("solid", fgColor="DDEEFF")
GRAY_FILL     = PatternFill("solid", fgColor="E8E8E8")

WHITE_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
BLACK_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT  = Font(name="微软雅黑", bold=True, size=10)
GRAY_FONT  = Font(name="微软雅黑", size=9, color="595959")
EXAMPLE_FONT = Font(name="微软雅黑", size=10, color="808080", italic=True)

THIN   = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)


def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:   c.font = font
    if fill:   c.fill = fill
    if align:  c.alignment = align
    if border: c.border = border
    return c


def build_sheet(wb, title, headers, example_row, validations, info_text):
    """
    headers: list of (字段名, 说明, 列宽, 必填bool)
    validations: list of (sqref, formula1)
    """
    ws = wb.create_sheet(title) if title else wb.active
    if not title:
        ws.title = "子租户主档"

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22

    # 说明行
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col}1")
    set_cell(ws, 1, 1, info_text,
             font=GRAY_FONT, fill=INFO_FILL, align=LEFT, border=BORDER)

    # 字段名行 + 说明行
    for col_idx, (name, desc, width, required) in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        set_cell(ws, 2, col_idx, name,
                 font=WHITE_FONT, fill=HEADER_FILL, align=CENTER, border=BORDER)
        set_cell(ws, 3, col_idx, desc,
                 font=GRAY_FONT,
                 fill=REQUIRED_FILL if required else OPTIONAL_FILL,
                 align=LEFT, border=BORDER)

    # 示例行
    for col_idx, val in enumerate(example_row, start=1):
        set_cell(ws, 4, col_idx, val,
                 font=EXAMPLE_FONT, fill=EXAMPLE_FILL,
                 align=LEFT, border=BORDER)

    # 数据验证
    for sqref, formula in validations:
        dv = DataValidation(
            type="list", formula1=formula, allow_blank=False,
            showErrorMessage=True, error="请从列表中选择", errorTitle="输入错误"
        )
        ws.add_data_validation(dv)
        dv.sqref = sqref

    return ws


wb = openpyxl.Workbook()

# ──────────────────────────────────────────────
# Sheet 1：子租户主档
# ──────────────────────────────────────────────
headers1 = [
    ("所属主合同编号★",  "与03_合同清单中「合同编号」完全一致，必填",        24, True),
    ("二房东企业名称★",  "主合同签订方企业全称",                            24, True),
    ("终端租客名称★",    "企业填全称；个人填姓名",                          20, True),
    ("租客类型★",        "企业 / 个人",                                    12, True),
    ("统一社会信用代码", "企业租客必填；个人租客留空",                       26, False),
    ("身份证号",          "个人租客必填；系统加密存储，脱敏后4位",          22, False),
    ("联系人姓名★",      "企业填对接负责人；个人即本人",                    16, True),
    ("联系电话★",        "系统加密存储，脱敏后4位",                        16, True),
    ("紧急联系方式",      "备用联系人姓名+电话",                            22, False),
    ("关联单元编号★",    "实际使用单元，多个用英文逗号分隔",                26, True),
    ("入住状态★",        "已入住/已签未入住/已退租/空置",                   18, True),
    ("备注",              "其他需说明情况",                                 20, False),
]

example1 = [
    "CONTRACT-2024-001", "北京科创有限公司", "上海贸易（租客）有限公司",
    "企业", "91110000123456789X", "",
    "张三", "13800138000", "李四 13900139000",
    "A-5F-01,A-5F-02", "已入住", "",
]

validations1 = [
    ("D4:D1000", '"企业,个人"'),
    ("K4:K1000", '"已入住,已签未入住,已退租,空置"'),
]

info1 = (
    "【填写说明】★=必填 ☆=建议填写。本表每行代表一位终端租客。"
    "所属主合同编号须与03_合同清单编号完全一致。"
    "联系电话/证件号系统加密存储，仅展示后4位。"
    "第4行为示例，提交前请删除。"
)

build_sheet(wb, None, headers1, example1, validations1, info1)

# ──────────────────────────────────────────────
# Sheet 2：子租赁合同
# ──────────────────────────────────────────────
headers2 = [
    ("所属主合同编号★",   "与03_合同清单编号完全一致",                     24, True),
    ("子租赁合同编号★",   "无正式编号可填「主合同号-01」形式",              24, True),
    ("终端租客名称★",     "须与子租户主档名称完全一致",                    20, True),
    ("关联单元编号★",     "实际占用单元，多个用英文逗号分隔",              26, True),
    ("计费面积(m²)★",     "合同约定计费面积，纯数字",                      14, True),
    ("起租日★",            "YYYY-MM-DD",                                   14, True),
    ("到期日★",            "YYYY-MM-DD",                                   14, True),
    ("月租金(元)★",        "含税月总租金，纯数字",                         14, True),
    ("含税/不含税★",       "含税 / 不含税",                                14, True),
    ("适用税率(%)",         "如 9 或 5；不含税时必填",                      12, False),
    ("押金金额(元)★",      "纯数字，0表示无押金",                          14, True),
    ("付款周期★",          "月付 / 季付 / 年付",                           12, True),
    ("免租期起始日",        "YYYY-MM-DD，无免租期留空",                     16, False),
    ("免租期结束日",        "YYYY-MM-DD，无免租期留空",                     16, False),
    ("备注",                "如有递增条款请说明或另附",                     22, False),
]

example2 = [
    "CONTRACT-2024-001", "CONTRACT-2024-001-01",
    "上海贸易（租客）有限公司", "A-5F-01,A-5F-02",
    "280.00", "2024-03-01", "2026-02-28",
    "28000", "含税", "9",
    "84000", "季付",
    "2024-03-01", "2024-03-31", "免租1个月"
]

validations2 = [
    ("I4:I1000", '"含税,不含税"'),
    ("L4:L1000", '"月付,季付,年付"'),
]

info2 = (
    "【填写说明】★=必填 ☆=建议填写。本表每行代表一份子租赁合同。"
    "需与「子租户主档」通过「终端租客名称」+「所属主合同编号」关联。"
    "日期格式统一 YYYY-MM-DD。第4行为示例，提交前请删除。"
)

build_sheet(wb, "子租赁合同", headers2, example2, validations2, info2)

# ──────────────────────────────────────────────
# Sheet 3：说明与图例
# ──────────────────────────────────────────────
ws3 = wb.create_sheet("说明与图例")
ws3.column_dimensions["A"].width = 22
ws3.column_dimensions["B"].width = 55

ws3.merge_cells("A1:B1")
set_cell(ws3, 1, 1, "PropOS 子租赁数据导入模板说明",
         font=WHITE_FONT, fill=HEADER_FILL, align=CENTER, border=BORDER)
ws3.row_dimensions[1].height = 28

notes = [
    ("模板版本",    "v1.0 · 2026-04-28",                                       False),
    ("适用模块",    "M5 二房东穿透管理",                                        False),
    ("Sheet 1",    "子租户主档 — 每位终端租客一行",                             False),
    ("Sheet 2",    "子租赁合同 — 每份子租赁合同一行",                           False),
    ("",           "",                                                           False),
    ("颜色说明",   "",                                                           True),
    ("■ 深蓝表头", "字段名称行",                                                False),
    ("■ 浅黄底色", "★ 必填字段（留空将导致导入失败）",                          False),
    ("■ 浅绿底色", "☆ 建议填写字段（可留空，建议补全）",                        False),
    ("■ 灰色斜体", "示例数据行（请替换实际数据后删除）",                        False),
    ("",           "",                                                           False),
    ("注意事项",   "",                                                           True),
    ("1",          "日期格式统一 YYYY-MM-DD，如 2024-03-01",                    False),
    ("2",          "联系电话与证件号系统加密存储，仅展示后4位",                 False),
    ("3",          "所属主合同编号须与 03_合同清单 文件中保持完全一致",         False),
    ("4",          "多个单元编号以英文逗号分隔，如：A-5F-01,A-5F-02",          False),
    ("5",          "填写完成后请删除第4行示例行再提交",                         False),
    ("6",          "子租赁续签关系请在备注列注明原合同编号",                    False),
    ("7",          "同一终端租客占用多个单元但为独立计费时，各行单独填写",      False),
]

for r_idx, (k, v, is_section) in enumerate(notes, start=2):
    ws3.row_dimensions[r_idx].height = 18
    set_cell(ws3, r_idx, 1, k,
             font=BOLD_FONT if is_section else BLACK_FONT,
             fill=GRAY_FILL if is_section else None,
             align=LEFT, border=BORDER)
    set_cell(ws3, r_idx, 2, v,
             font=BLACK_FONT,
             fill=GRAY_FILL if is_section else None,
             align=LEFT, border=BORDER)

out = "/Users/sheltonwan/MyApps/propos/docs/templates/kickoff/10_子租赁数据导入_租客与合同.xlsx"
wb.save(out)
print(f"已生成：{out}")
