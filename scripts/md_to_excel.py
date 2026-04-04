#!/usr/bin/env python3
"""将任意 Markdown 文件中的表格导出为 Excel（每个表格一个 Sheet）。

用法:
    python3 md_to_excel.py <input.md> [output.xlsx]

如果不指定 output.xlsx，则自动在同目录下生成同名 .xlsx 文件。
"""

import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── helpers ──────────────────────────────────────────────────────────

def parse_md_table(lines: list[str]) -> tuple[list[str], list[list[str]]]:
    """Parse a markdown table block into (headers, rows)."""
    headers = [c.strip() for c in lines[0].strip().strip("|").split("|")]
    rows = []
    for line in lines[2:]:  # skip separator line
        cells = [c.strip() for c in line.strip().strip("|").split("|")]
        rows.append(cells)
    return headers, rows


def find_tables(text: str) -> list[tuple[str, list[str], list[list[str]]]]:
    """Find all tables in markdown text, return (preceding_heading, headers, rows)."""
    lines = text.split("\n")
    tables: list[tuple[str, list[str], list[list[str]]]] = []
    current_heading = ""
    i = 0
    while i < len(lines):
        line = lines[i]
        # Track headings
        m = re.match(r"^#{1,5}\s+(.+)", line)
        if m:
            current_heading = m.group(1).strip()
        # Detect table start: line with | and next line is separator
        if "|" in line and i + 1 < len(lines) and re.match(r"^\s*\|[\s:|-]+\|\s*$", lines[i + 1]):
            table_lines = [line]
            j = i + 1
            while j < len(lines) and "|" in lines[j] and lines[j].strip():
                table_lines.append(lines[j])
                j += 1
            if len(table_lines) >= 3:
                headers, rows = parse_md_table(table_lines)
                tables.append((current_heading, headers, rows))
            i = j
            continue
        i += 1
    return tables


def clean_cell(val: str) -> str:
    """Remove markdown formatting for Excel display."""
    val = re.sub(r"\*\*(.+?)\*\*", r"\1", val)  # bold
    val = re.sub(r"\[(.+?)\]\(.+?\)", r"\1", val)  # links
    val = re.sub(r"`(.+?)`", r"\1", val)  # inline code
    return val.strip()


def auto_width(ws):
    """Auto-adjust column widths."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    # Approximate: CJK chars count as 2
                    text = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in text)
                    max_len = max(max_len, length)
        adjusted = min(max_len + 4, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


# ── style constants ──────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def make_sheet_name(heading: str, index: int, seen: dict[str, int]) -> str:
    """Generate a unique Excel sheet name (max 31 chars) from heading."""
    # Remove emoji and special chars for cleaner sheet name
    name = re.sub(r"[🔴🟡🟢⚪✅⬜🔶⚙️➕➖]", "", heading).strip()
    # Remove leading numbering like "一、" "二、" "1." "1)"
    name = re.sub(r"^[一二三四五六七八九十\d]+[、.)\s]+", "", name).strip()
    if not name:
        name = f"表格{index + 1}"
    # Truncate to 31 chars (Excel limit)
    name = name[:31]
    # Ensure uniqueness
    if name in seen:
        seen[name] += 1
        suffix = f"({seen[name]})"
        name = name[:31 - len(suffix)] + suffix
    else:
        seen[name] = 1
    return name


def write_table_to_sheet(ws, headers: list[str], rows: list[list[str]]):
    """Write a parsed table to a worksheet with styling."""
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=clean_cell(h))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=clean_cell(val))
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER
    auto_width(ws)
    ws.freeze_panes = "A2"


def main():
    if len(sys.argv) < 2:
        print(f"用法: python3 {Path(__file__).name} <input.md> [output.xlsx]")
        sys.exit(1)

    md_file = Path(sys.argv[1]).resolve()
    if not md_file.exists():
        print(f"❌ 文件不存在: {md_file}")
        sys.exit(1)

    if len(sys.argv) >= 3:
        out_file = Path(sys.argv[2]).resolve()
    else:
        out_file = md_file.with_suffix(".xlsx")

    text = md_file.read_text(encoding="utf-8")
    tables = find_tables(text)

    if not tables:
        print(f"⚠️ 未在 {md_file.name} 中找到任何 Markdown 表格")
        sys.exit(0)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    seen_names: dict[str, int] = {}
    for idx, (heading, headers, rows) in enumerate(tables):
        sheet_name = make_sheet_name(heading, idx, seen_names)
        ws = wb.create_sheet(title=sheet_name)
        write_table_to_sheet(ws, headers, rows)

    wb.save(out_file)
    print(f"✅ 导出完成: {out_file}")
    print(f"   来源: {md_file.name}")
    print(f"   共 {len(wb.sheetnames)} 个 Sheet: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
