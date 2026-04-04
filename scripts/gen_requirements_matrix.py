#!/usr/bin/env python3
"""生成 PropOS 全量需求清单 Excel（Phase 1 + Phase 2 + Phase 3）。

Phase 1 & 2 来源：PRD.md
Phase 3 来源：COMPETITIVE_ANALYSIS.md 中建议新增但 PRD 未覆盖的需求
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_FILE = Path(__file__).resolve().parent.parent / "docs" / "PropOS_需求全景清单.xlsx"

# ── 成本基准（外包市场行情 2025-2026）──────────────────────────────
DAY_COST_T1 = 1500          # 一线城市中级日薪（北上广深杭）
DAY_COST_T1_HIGH = 2500     # 一线城市高级日薪（北上广深杭）
DAY_COST_T2 = 1000          # 二线城市中级日薪（成都/武汉/长沙/西安）
DAY_COST = DAY_COST_T1      # 兼容保留，默认一线中级

# ── 样式常量 ──────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
PHASE_FILLS = {
    "Phase 1": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Phase 2": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "Phase 3": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}
PRIORITY_FILLS = {
    "P0": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "P1": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "P2": PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
    "P3": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

# ── 需求数据 ──────────────────────────────────────────────────────────
# 格式: (阶段, 模块, 编号, 名称, 描述, 优先级, 工期天数, 来源, 备注)

REQUIREMENTS = [
    # ═══════════════════════════════════════════════════════════════════
    # Phase 1 — M0: 基础架构（4 周）
    # ═══════════════════════════════════════════════════════════════════
    ("Phase 1", "M0-基础架构", "M0.1", "系统架构设计", "技术选型、微服务/单体架构决策、API 规范制定", "P0", 5, "PRD", ""),
    ("Phase 1", "M0-基础架构", "M0.2", "数据库设计", "ER 模型设计、字段定义、索引策略、多租户方案", "P0", 5, "PRD", ""),
    ("Phase 1", "M0-基础架构", "M0.3", "权限与角色模块", "RBAC 角色权限体系：超级管理员/运营管理层/租务专员/财务/维修主管/二房东", "P0", 5, "PRD", ""),
    ("Phase 1", "M0-基础架构", "M0.4", "UI/UX 设计", "设计系统、组件库、交互原型、视觉稿", "P0", 5, "PRD", ""),

    # ═══════════════════════════════════════════════════════════════════
    # Phase 1 — M1: 资产与空间可视化（5 周）
    # ═══════════════════════════════════════════════════════════════════
    ("Phase 1", "M1-资产与空间可视化", "M1.1", "楼栋档案管理", "楼栋名称、楼层总数、业态类型、GFA/NLA、多楼栋并行管理", "P0", 3, "PRD", ""),
    ("Phase 1", "M1-资产与空间可视化", "M1.2", "CAD 平面图导入与展示", "支持 .dwg 转换为 SVG/PNG 分层展示，按楼层切换浏览，系统在图层上叠加单元热区与状态色块", "P0", 5, "PRD", "技术风险：CAD 转换兼容性"),
    ("Phase 1", "M1-资产与空间可视化", "M1.3", "单元元数据档案", "单元编号、套内面积、层高、朝向、装修状态、业态分类、改造历史、出租状态", "P0", 3, "PRD", ""),
    ("Phase 1", "M1-资产与空间可视化", "M1.4", "业态差异化字段", "写字楼：工位数/分隔间数；商铺：门面宽度/临街面/层高；公寓：卧室数/独立卫生间", "P0", 2, "PRD", ""),
    ("Phase 1", "M1-资产与空间可视化", "M1.5", "楼层切片状态色块", "已租(绿)/即将到期(黄)/空置(红)/非可租区域(灰)，按业态筛选过滤", "P0", 3, "PRD", ""),
    ("Phase 1", "M1-资产与空间可视化", "M1.6", "改造记录管理", "绑定单元，记录改造类型/日期/施工造价，支持改造前后照片上传", "P1", 2, "PRD", ""),
    ("Phase 1", "M1-资产与空间可视化", "M1.7", "资产台账导出", "全部 639 套单元数据按业态分类导出 Excel", "P0", 2, "PRD", ""),
    ("Phase 1", "M1-资产与空间可视化", "M1.8", "资产概览看板", "按业态汇总：总套数/已租/空置/出租率（写字楼/商铺/公寓三列）", "P0", 3, "PRD", ""),

    # ═══════════════════════════════════════════════════════════════════
    # Phase 1 — M2: 租务与合同管理（6 周）
    # ═══════════════════════════════════════════════════════════════════
    # -- 2.1 租客全景画像 --
    ("Phase 1", "M2-租务与合同管理", "M2.1.1", "租客基本信息", "企业/个人、统一社会信用代码/身份证、联系人、紧急联系方式", "P0", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.1.2", "租赁历史追踪", "当前及历史合同、曾租单元、续租次数", "P0", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.1.3", "缴费信用评记", "历史逾期次数、最近付款日、系统自动生成信用评级(A/B/C)", "P1", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.1.4", "工单关联", "关联该租客历史报修记录", "P1", 1, "PRD", ""),
    # -- 2.2 合同管理 --
    ("Phase 1", "M2-租务与合同管理", "M2.2.1", "合同录入", "单元绑定、起租日、到期日、月租金、免租/装修期、押金、付款周期、业态自动关联", "P0", 3, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.2.2", "附件管理", "合同 PDF 及补充协议上传", "P0", 1, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.2.3", "合同状态机", "报价中→待签约→执行中→即将到期→已到期/续签/已终止", "P0", 3, "PRD", "竞品建议裁剪'报价中'阶段(C07)"),
    ("Phase 1", "M2-租务与合同管理", "M2.2.4", "续签管理", "续签合同与原合同关联，形成完整合同链", "P0", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.2.5", "免租/装修期处理", "免租期内账单自动标记，不计入逾期", "P0", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.2.6", "商铺营业额分成", "保底租金+营业额分成比例，按月录入营业额自动计算实收租金", "P2", 3, "PRD", "竞品建议简化(C06)：仅备注字段"),
    # -- 2.3 智能预警引擎 --
    ("Phase 1", "M2-租务与合同管理", "M2.3.1", "租约到期预警", "提前 90/60/30 天触发通知", "P0", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.3.2", "租金逾期预警", "账期截止后第 1/7/15 天未到账触发通知", "P0", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.3.3", "月度到期汇总", "每月 1 日自动推送到期汇总给管理层", "P1", 1, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.3.4", "押金退还提醒", "合同终止前 7 天提醒财务", "P1", 1, "PRD", ""),
    # -- 2.4 动态 WALE 计算 --
    ("Phase 1", "M2-租务与合同管理", "M2.4.1", "组合级 WALE", "全部在租合同加权平均到期年数", "P0", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.4.2", "楼栋级 WALE", "按楼栋独立计算", "P0", 1, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.4.3", "业态级 WALE", "写字楼/商铺/公寓分别独立呈现", "P0", 1, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.4.4", "WALE 趋势图", "过去 12 个月变化曲线", "P1", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.4.5", "到期瀑布图", "按年份显示未来租约到期面积分布", "P1", 2, "PRD", ""),
    # -- 2.5 租金递增规则配置器 --
    ("Phase 1", "M2-租务与合同管理", "M2.5.1", "固定比例递增", "每个递增周期按固定百分比上涨", "P0", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.5.2", "固定金额递增", "每个递增周期按固定单价上涨", "P0", 1, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.5.3", "阶梯式递增", "按合同年限分段设定不同租金", "P0", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.5.4", "CPI 挂钩递增", "每年按上年度 CPI 涨幅调整（手动录入）", "P1", 2, "PRD", "竞品建议简化(C02)：去掉统计局API对接"),
    ("Phase 1", "M2-租务与合同管理", "M2.5.5", "每 N 年递增", "指定递增间隔年数与涨幅", "P2", 1, "PRD", "竞品建议可合并到固定比例递增(C03)"),
    ("Phase 1", "M2-租务与合同管理", "M2.5.6", "免租后基准调整", "免租期结束后，首年按约定基准价起算", "P2", 1, "PRD", "竞品建议可合并实现(C03)"),
    ("Phase 1", "M2-租务与合同管理", "M2.5.7", "递增模板管理", "创建、保存、套用递增模板，按业态分类命名", "P1", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.5.8", "默认模板", "为每个业态设定默认递增模板", "P1", 1, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.5.9", "自动账单联动", "递增规则自动计算每期应收租金，无需人工修改", "P0", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.5.10", "WALE 计算联动", "WALE 公式取递增后实际租金", "P0", 1, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.5.11", "未来租金预测", "自动生成合同全生命周期租金预测表，支持导出", "P1", 2, "PRD", ""),
    ("Phase 1", "M2-租务与合同管理", "M2.5.12", "续签对比", "自动对比原合同末期租金与新合同起始租金，计算涨跌幅", "P1", 1, "PRD", ""),

    # ═══════════════════════════════════════════════════════════════════
    # Phase 1 — M3: 财务与业财一体化（5 周）
    # ═══════════════════════════════════════════════════════════════════
    # -- 3.1 账单与收款 --
    ("Phase 1", "M3-财务与业财一体化", "M3.1.1", "自动账单生成", "依据合同条款，在每个账期前自动生成应收账单", "P0", 3, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.1.2", "多费项支持", "租金、物业管理费（各业态费率独立）、水电代收、停车费、储藏室", "P0", 2, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.1.3", "收款核销", "财务录入到账信息，系统自动匹配账单并标记核销", "P0", 3, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.1.4", "发票管理", "记录已开票/未开票状态，录入发票号", "P1", 2, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.1.5", "自动催收", "逾期后按节点向租客发送催收提醒（邮件/短信模板可配置）", "P0", 2, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.1.6", "账单导出", "按时间段/楼栋/业态/租客维度导出 Excel", "P0", 1, "PRD", ""),
    # -- 3.2 NOI 实时看板 --
    ("Phase 1", "M3-财务与业财一体化", "M3.2.1", "月度 NOI 卡片", "当月实收租金/运营支出/NOI 三栏并列实时展示", "P0", 2, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.2.2", "业态 NOI 分拆", "写字楼/商铺/公寓三业态分别展示收入、支出与 NOI 回报率", "P0", 2, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.2.3", "运营支出类目管理", "水电公摊、保洁保安、维修费、保险、税金", "P0", 2, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.2.4", "出租率看板", "已租 NLA / 总可租 NLA，可按业态下钻", "P0", 1, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.2.5", "空置损失测算", "空置单元的市值租金损失估算", "P1", 2, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.2.6", "收款进度", "本月应收 vs 实收对比，快速检索未缴款租户", "P0", 2, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.2.7", "NOI 历史趋势", "过去 12 个月 NOI 折线图", "P1", 2, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.2.8", "楼栋下钻", "按楼栋独立查看 NOI", "P1", 1, "PRD", ""),
    # -- 3.3 KPI 仪表盘 --
    ("Phase 1", "M3-财务与业财一体化", "M3.3.1", "KPI 方案配置", "方案名称、指标选择、权重分配、阈值调整、评估周期、适用对象、生效时间", "P1", 5, "PRD", "竞品建议降级(C01)：改为预定义指标看板"),
    ("Phase 1", "M3-财务与业财一体化", "M3.3.2", "KPI 自动打分", "10 个预定义 KPI 指标自动从各模块抽取数据计算得分", "P0", 3, "PRD", ""),
    ("Phase 1", "M3-财务与业财一体化", "M3.3.3", "KPI 总览看板", "按部门/员工展示当期 KPI 总分、各指标得分雷达图", "P1", 3, "PRD", "竞品建议降级(C01)：去掉雷达图和排名"),
    ("Phase 1", "M3-财务与业财一体化", "M3.3.4", "KPI 排名榜", "同一 KPI 方案下的员工/部门排名", "P2", 2, "PRD", "竞品建议裁剪(C01)"),
    ("Phase 1", "M3-财务与业财一体化", "M3.3.5", "KPI 历史趋势", "过去 6~12 个月 KPI 得分折线图", "P2", 2, "PRD", "竞品建议裁剪(C01)"),
    ("Phase 1", "M3-财务与业财一体化", "M3.3.6", "KPI 导出报告", "KPI 评分结果导出 PDF/Excel", "P2", 1, "PRD", ""),

    # ═══════════════════════════════════════════════════════════════════
    # Phase 1 — M4: 物业运营与工单（4 周）
    # ═══════════════════════════════════════════════════════════════════
    ("Phase 1", "M4-物业运营与工单", "M4.1.1", "Flutter App 报修", "选择楼栋→楼层→单元→问题类型，照片上传(最多5张)，紧急程度标记", "P0", 5, "PRD", ""),
    ("Phase 1", "M4-物业运营与工单", "M4.1.2", "微信小程序报修(精简版)", "扫码报修+查看状态，照片上传(最多3张)", "P2", 4, "PRD", "竞品建议延至 Phase 1.5(C04)"),
    ("Phase 1", "M4-物业运营与工单", "M4.1.3", "工单状态实时追踪", "提报人可实时查看工单进度", "P0", 2, "PRD", ""),
    ("Phase 1", "M4-物业运营与工单", "M4.1.4", "工单推送通知", "APNs/FCM 原生通知栏推送", "P0", 2, "PRD", ""),
    ("Phase 1", "M4-物业运营与工单", "M4.1.5", "扫码报修", "扫单元二维码自动定位", "P0", 1, "PRD", ""),
    ("Phase 1", "M4-物业运营与工单", "M4.1.6", "CAD 楼层快查(App)", "矢量渲染+缩放平移", "P1", 2, "PRD", ""),
    ("Phase 1", "M4-物业运营与工单", "M4.2", "工单流转状态机", "已提交→已审核/派单→处理中→待验收→已完成（含拒绝/挂起分支）", "P0", 3, "PRD", ""),
    ("Phase 1", "M4-物业运营与工单", "M4.3.1", "维修费用录入", "内部人员完工后录入材料费+人工费", "P0", 1, "PRD", ""),
    ("Phase 1", "M4-物业运营与工单", "M4.3.2", "成本归口", "费用关联具体单元/楼层/楼栋，自动汇入 NOI 运营支出", "P0", 2, "PRD", ""),
    ("Phase 1", "M4-物业运营与工单", "M4.3.3", "维修成本报表", "按时间段/楼栋/费用类型查看维修费用汇总", "P1", 2, "PRD", ""),
    ("Phase 1", "M4-物业运营与工单", "M4.3.4", "供应商管理", "维护常用维修供应商信息", "P1", 1, "PRD", ""),

    # ═══════════════════════════════════════════════════════════════════
    # Phase 1 — M5: 二房东穿透管理（4 周）
    # ═══════════════════════════════════════════════════════════════════
    ("Phase 1", "M5-二房东穿透管理", "M5.1", "主合同-子租赁两级数据模型", "一个主合同(Master Lease)下可关联 N 条子租赁(Sub-Lease)", "P0", 2, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.2", "子租赁数据字段", "终端租客名称/类型/联系方式/证件/起止日/月租金/单价/入住状态等 12 个字段", "P0", 2, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.1", "内部手工录入", "租务专员在管理后台逐条录入子租赁", "P0", 1, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.2", "Excel 批量导入", "提供子租赁导入模板，支持按二房东批量导入", "P0", 2, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.3", "数据校验", "单元编号必须在主合同范围内；子租赁到期日不超主合同；单元不可重复", "P0", 1, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.4", "二房东独立登录入口", "为每家二房东分配独立账号，Web 端轻量表单页面", "P0", 3, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.5", "数据权限隔离", "二房东仅可查看/编辑自身主合同范围内的单元", "P0", 2, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.6", "二房东填报表单", "二房东登录后可对名下每单元填写/更新租客信息", "P0", 2, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.7", "二房东 Excel 上传", "二房东可下载标准模板，填写后批量上传", "P1", 1, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.8", "变更记录留痕", "记录操作时间、修改前后内容、操作人", "P0", 1, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.9", "数据审核机制", "二房东提交为待审核，审核通过后方生效", "P0", 2, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.10", "提交提醒", "每月固定时间自动向二房东发送填报提醒", "P1", 1, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.3.11", "二房东端 API 预留", "预留 RESTful API 接口文档", "P2", 2, "PRD", "竞品建议裁剪(C05)：YAGNI"),
    ("Phase 1", "M5-二房东穿透管理", "M5.4.1", "二房东总览卡片", "主合同租金/已填报子租赁数/终端出租面积/终端空置面积", "P0", 2, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.4.2", "转租溢价分析", "终端平均租金单价 vs 主合同租金单价，计算溢价率", "P0", 1, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.4.3", "穿透出租率", "二房东名下单元的终端实际出租率", "P0", 1, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.4.4", "子租赁到期预警", "终端租客合同集中到期预警", "P1", 1, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.4.5", "填报完整度监控", "监控二房东数据填报覆盖率", "P1", 1, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.4.6", "楼层穿透视图", "在楼层色块图上叠加子租赁信息", "P1", 2, "PRD", ""),
    ("Phase 1", "M5-二房东穿透管理", "M5.5", "安全与权限设计", "账号管理、数据隔离、操作审计、接口安全、数据脱敏、自动冻结", "P0", 3, "PRD", ""),

    # ═══════════════════════════════════════════════════════════════════
    # Phase 1 — M9: 测试与上线（2 周）
    # ═══════════════════════════════════════════════════════════════════
    ("Phase 1", "M9-测试与上线", "M9.1", "集成测试", "全模块集成测试、边界场景验证", "P0", 5, "PRD", ""),
    ("Phase 1", "M9-测试与上线", "M9.2", "UAT 用户验收测试", "关键用户参与验收，验收标准逐项确认", "P0", 3, "PRD", ""),
    ("Phase 1", "M9-测试与上线", "M9.3", "数据初始化", "639 套单元导入、500+ 在租合同导入、运营支出初始化", "P0", 5, "PRD", "约 3~4 周，与 UAT 并行"),
    ("Phase 1", "M9-测试与上线", "M9.4", "培训与上线", "用户培训、上线切换、试运行支持", "P0", 2, "PRD", ""),

    # ═══════════════════════════════════════════════════════════════════
    # Phase 2 — PRD 已规划（参考级）
    # ═══════════════════════════════════════════════════════════════════
    ("Phase 2", "M6-租户服务", "M6.1", "租户自助门户", "查账单、在线缴租、报修、查合同", "P0", 35, "PRD", "¥8万~15万"),
    ("Phase 2", "M6-租户服务", "M6.2", "在线付款中心", "对接微信/支付宝，租户在线缴费", "P0", 18, "PRD", "¥4万~8万"),
    ("Phase 2", "M6-租户服务", "M6.3", "电子合同与签章", "集成法大大/e签宝，无纸化签署", "P1", 20, "PRD", "¥5万~10万"),
    ("Phase 2", "M7-招商与出租", "M7.1", "在线出租模块", "空置单元展示、预约看房、在线报价", "P1", 35, "PRD", "¥10万~18万"),
    ("Phase 2", "M7-招商与出租", "M7.2", "招商 CRM", "线索管理、跟进记录、转化漏斗", "P1", 0, "PRD", "含在出租模块工期内"),
    ("Phase 2", "M8-物业扩展", "M8.1", "智能门禁集成", "对接海康/华为云鼎，按合同自动授权/撤权", "P0", 25, "PRD", "¥6万~12万"),
    ("Phase 2", "M8-物业扩展", "M8.2", "外包物业工作台", "工单接收、费用单、日常巡检、设备保养计划", "P0", 25, "PRD", "¥6万~12万"),

    # ═══════════════════════════════════════════════════════════════════
    # Phase 3 — 竞品分析建议新增
    # ═══════════════════════════════════════════════════════════════════
    ("Phase 3", "竞品建议-P0", "A01", "入退租流程管理", "结构化入住/退租检查清单工作流：验房→钥匙交接→水电起度→押金确认→退租验房→损坏扣款→退押金→账号注销，每步拍照留痕", "P0", 8, "竞品分析", "6/6 竞品覆盖，建议纳入 Phase 1"),
    ("Phase 3", "竞品建议-P0", "A02", "通知/公告中心", "系统内统一消息中心：系统通知集中展示、管理层发布公告、已读/未读统计。预警/工单/催收的统一出口", "P0", 5, "竞品分析", "5/6 竞品覆盖，建议纳入 Phase 1"),
    ("Phase 3", "竞品建议-P0", "A03", "标准租金卷报表(Rent Roll)", "标准化汇总报表：每单元的租客名称、合同起止日、月租金、面积、单价、付款周期、下次到期日。支持筛选和导出", "P0", 3, "竞品分析", "5/6 竞品覆盖，业主/投资人/银行必看"),
    ("Phase 3", "竞品建议-P0", "A04", "年度预算管理与差异分析", "按楼栋/业态设定年度收支预算，NOI 看板增加预算 vs 实际对比，月度差异自动标红。KPI K07 的必要前置", "P0", 4, "竞品分析", "4/6 竞品覆盖，K07 指标无此无法落地"),
    ("Phase 3", "竞品建议-P1", "A05", "能耗/水电抄表管理", "水电燃气分表台账、周期性抄表记录（人工录入或照片识别）、自动生成水电费账单并关联租客", "P1", 8, "竞品分析", "4/6 竞品覆盖"),
    ("Phase 3", "竞品建议-P1", "A06", "设备资产台账与保养计划", "电梯/空调/消防等大型设备档案、定期保养计划自动触发工单、设备报修历史统计", "P1", 8, "竞品分析", "4/6 竞品覆盖"),
    ("Phase 3", "竞品建议-P1", "A07", "文档管理中心", "按楼栋/单元/合同组织文档树，支持上传各类文档，文档到期提醒，全文搜索", "P1", 5, "竞品分析", "3/6 竞品覆盖"),
    ("Phase 3", "竞品建议-P1", "A08", "停车场管理", "车位台账、车位绑定租户合同、停车费账单自动生成、停车收入纳入 NOI", "P1", 4, "竞品分析", "3/6 竞品覆盖（国内普遍）"),
    ("Phase 3", "竞品建议-P1", "A09", "访客管理", "来访人员登记（姓名/证件/事由/被访人）、生成访客二维码凭证、访客记录查询", "P1", 4, "竞品分析", "3/6 竞品覆盖，合规基础组件"),
    ("Phase 3", "竞品建议-P2", "A10", "市场租金对标分析", "录入周边同类物业挂牌租金数据，本物业 vs 市场均价对比，续签/新签定价参考", "P2", 8, "竞品分析", "3/6 竞品覆盖"),
    ("Phase 3", "竞品建议-P2", "A11", "会议室/共享空间预订", "公共会议室在线预约、按时段/按次计费、预订冲突检测", "P2", 4, "竞品分析", "2/6 竞品覆盖"),
    ("Phase 3", "竞品建议-P2", "A12", "空置率预测(AI)", "基于历史出租/退租趋势+合同到期时间预测未来 3~6 个月空置率走势", "P2", 8, "竞品分析", "1/6 竞品覆盖，需积累历史数据"),
    ("Phase 3", "竞品建议-P2", "A13", "保险管理", "物业整体保险到期追踪、租户责任险购买记录、到期自动提醒", "P2", 3, "竞品分析", "3/6 竞品覆盖"),
    ("Phase 3", "竞品建议-P2", "A14", "巡检管理", "按楼栋/楼层设定巡检路线和检查项、NFC/二维码打卡巡检、异常项自动生成工单", "P2", 8, "竞品分析", "3/6 竞品覆盖，可纳入 Phase 2 外包物业工作台"),
    ("Phase 3", "竞品建议-P2", "A15", "Open API 平台", "标准 RESTful API + 开发者文档，支持第三方系统集成（ERP/OA/财务软件）", "P2", 5, "竞品分析", "3/6 竞品覆盖"),
    ("Phase 3", "竞品建议-P3", "A16", "租户背调/信用筛查", "对接征信平台自动查询企业/个人信用", "P3", 0, "竞品分析", "3/6 竞品覆盖，规模>2000套时考虑"),
    ("Phase 3", "竞品建议-P3", "A17", "多语言支持", "系统支持中英文切换", "P3", 0, "竞品分析", "2/6 竞品覆盖，引入外资租户时考虑"),
    ("Phase 3", "竞品建议-P3", "A18", "物业费催缴+诉讼流程管理", "催缴信函自动生成、律师函模板、诉讼进度跟踪", "P3", 0, "竞品分析", "0/6 竞品覆盖，规模>5000套时考虑"),
    ("Phase 3", "竞品建议-P3", "A19", "BI 自助分析平台", "拖拽式自定义报表/仪表盘生成器", "P3", 0, "竞品分析", "0/6 竞品覆盖，可对接 Metabase/Superset"),
    ("Phase 3", "竞品建议-P3", "A20", "IoT 设备接入平台", "智能电表/水表/温湿度传感器实时数据采集", "P3", 0, "竞品分析", "0/6 竞品覆盖，需硬件改造投入"),
]

COLUMNS = [
    ("序号", 6), ("阶段", 12), ("模块", 22), ("需求编号", 10), ("需求名称", 28),
    ("需求描述", 55), ("优先级", 8), ("工期(人天)", 10), ("工期(人周)", 10),
    ("一线成本\n(中级·元)", 14), ("一线成本\n(高级·元)", 14), ("需求来源", 10), ("备注", 40),
]


def main():
    wb = Workbook()

    # ── Sheet 1: 全量需求清单 ──────────────────────────────────────
    ws = wb.active
    ws.title = "全量需求清单"

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    CITY_FILL_T1_COL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    CITY_FILL_T1_HIGH_COL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    for r_idx, (phase, module, req_id, name, desc, priority, days, source, note) in enumerate(REQUIREMENTS, 2):
        seq = r_idx - 1  # 序号从 1 开始
        weeks = round(days / 5, 1) if days > 0 else "待定"
        cost_t1 = round(days * DAY_COST_T1) if days > 0 else "待定"
        cost_t1_high = round(days * DAY_COST_T1_HIGH) if days > 0 else "待定"
        row_data = [seq, phase, module, req_id, name, desc, priority, days if days > 0 else "待定", weeks, cost_t1, cost_t1_high, source, note]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CELL_ALIGN if c_idx in (5, 6, 13) else CENTER_ALIGN
            if c_idx == 2 and phase in PHASE_FILLS:
                cell.fill = PHASE_FILLS[phase]
            if c_idx == 7 and priority in PRIORITY_FILLS:
                cell.fill = PRIORITY_FILLS[priority]
            if c_idx == 10 and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
                cell.fill = CITY_FILL_T1_COL
            if c_idx == 11 and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
                cell.fill = CITY_FILL_T1_HIGH_COL

    # 成本口径说明
    note_r = len(REQUIREMENTS) + 2
    note_text = (
        '※ 本表成本为“需求快速估算”口径：按统一日薪（中级 ¥1,500、高级 ¥2,500）× 需求人天计算，未区分角色。'
        '实际人员成本请参考“人员配置”Sheet（按角色各自日薪计算）。'
    )
    ws.cell(row=note_r, column=1, value=note_text).alignment = CELL_ALIGN
    ws.cell(row=note_r, column=1).font = Font(italic=True, color="666666")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: 阶段汇总 ────────────────────────────────────────
    ws2 = wb.create_sheet("阶段汇总")
    summary_headers = ["阶段", "需求数", "总工期(人天)", "总工期(人周)",
                       "一线成本\n(中级·元)", "一线成本\n(高级·元)", "说明"]
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    phase_stats = {}
    for phase, _, _, _, _, _, days, _, _ in REQUIREMENTS:
        if phase not in phase_stats:
            phase_stats[phase] = {"count": 0, "days": 0}
        phase_stats[phase]["count"] += 1
        phase_stats[phase]["days"] += days

    phase_notes = {
        "Phase 1": "PRD 五大模块 + 基础架构 + 测试上线（独立开发者 ~32 周）",
        "Phase 2": "PRD 已规划：租户门户、门禁、签章、在线出租、CRM、物业工作台、在线支付",
        "Phase 3": "竞品分析建议新增：20 项功能（P0×4 / P1×5 / P2×6 / P3×5）",
    }

    for r_idx, (phase, stats) in enumerate(phase_stats.items(), 2):
        weeks = round(stats["days"] / 5, 1)
        cost_t1 = round(stats["days"] * DAY_COST_T1)
        cost_t1_high = round(stats["days"] * DAY_COST_T1_HIGH)
        row_data = [phase, stats["count"], stats["days"], weeks, cost_t1, cost_t1_high, phase_notes.get(phase, "")]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if c_idx <= 6 else CELL_ALIGN
            if c_idx == 1 and phase in PHASE_FILLS:
                cell.fill = PHASE_FILLS[phase]
            if c_idx == 5:
                cell.fill = CITY_FILL_T1_COL
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
            if c_idx == 6:
                cell.fill = CITY_FILL_T1_HIGH_COL
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0"

    total_row = len(phase_stats) + 2
    total_count = sum(s["count"] for s in phase_stats.values())
    total_days = sum(s["days"] for s in phase_stats.values())
    total_data = ["合计", total_count, total_days, round(total_days / 5, 1),
                  round(total_days * DAY_COST_T1), round(total_days * DAY_COST_T1_HIGH), ""]
    for c_idx, val in enumerate(total_data, 1):
        cell = ws2.cell(row=total_row, column=c_idx, value=val)
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        cell.font = Font(bold=True)
        if c_idx == 5:
            cell.fill = CITY_FILL_T1_COL
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"
        if c_idx == 6:
            cell.fill = CITY_FILL_T1_HIGH_COL
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"

    # 成本说明行
    note_row = total_row + 1
    ws2.cell(row=note_row, column=1,
             value='※ “需求估算”口径：统一日薪（中级 ¥1,500、高级 ¥2,500）× 需求人天，未区分角色。实际人员成本请参考“人员配置”Sheet（按角色各自日薪计算）。').alignment = CELL_ALIGN
    ws2.cell(row=note_row, column=1).font = Font(italic=True, color="666666")

    for col_idx, w in enumerate([12, 10, 12, 12, 14, 14, 55], 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.freeze_panes = "A2"

    # ── Sheet 3: 模块汇总 ────────────────────────────────────────
    ws3 = wb.create_sheet("模块汇总")
    mod_headers = ["阶段", "模块", "需求数", "总工期(人天)", "总工期(人周)",
                   "一线成本\n(中级·元)", "一线成本\n(高级·元)"]
    for col_idx, h in enumerate(mod_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    module_stats = {}
    for phase, module, _, _, _, _, days, _, _ in REQUIREMENTS:
        key = f"{phase}|{module}"
        if key not in module_stats:
            module_stats[key] = {"phase": phase, "module": module, "count": 0, "days": 0}
        module_stats[key]["count"] += 1
        module_stats[key]["days"] += days

    for r_idx, (_, stats) in enumerate(module_stats.items(), 2):
        weeks = round(stats["days"] / 5, 1)
        cost_t1 = round(stats["days"] * DAY_COST_T1)
        cost_t1_high = round(stats["days"] * DAY_COST_T1_HIGH)
        row_data = [stats["phase"], stats["module"], stats["count"], stats["days"], weeks, cost_t1, cost_t1_high]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            if c_idx == 1 and stats["phase"] in PHASE_FILLS:
                cell.fill = PHASE_FILLS[stats["phase"]]
            if c_idx == 6:
                cell.fill = CITY_FILL_T1_COL
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
            if c_idx == 7:
                cell.fill = CITY_FILL_T1_HIGH_COL
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0"

    # 成本口径说明
    mod_note_row = len(module_stats) + 2
    ws3.cell(row=mod_note_row, column=1,
             value='※ “需求估算”口径：统一日薪 × 需求人天，未区分角色。实际人员成本请参考“人员配置”Sheet。').alignment = CELL_ALIGN
    ws3.cell(row=mod_note_row, column=1).font = Font(italic=True, color="666666")

    for col_idx, w in enumerate([12, 25, 10, 12, 12, 14, 14], 1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = w
    ws3.freeze_panes = "A2"

    # ── Sheet 4: 人员配置 ────────────────────────────────────────
    ws4 = wb.create_sheet("人员配置")
    staff_headers = [
        ("序号", 6), ("角色", 20), ("职责范围", 50), ("对应模块", 28),
        ("Phase 1\n人月", 10), ("Phase 2\n人月", 10), ("Phase 3\n人月", 10),
        ("合计\n人月", 9), ("建议来源", 12),
        ("初级\n日薪(元)", 10), ("初级\n月薪(元)", 11),
        ("中级\n日薪(元)", 10), ("中级\n月薪(元)", 11),
        ("高级\n日薪(元)", 10), ("高级\n月薪(元)", 11),
        ("初级\n总成本(元)", 12), ("中级\n总成本(元)", 12), ("高级\n总成本(元)", 12),
        ("备注", 40),
    ]
    for col_idx, (h, w) in enumerate(staff_headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws4.column_dimensions[get_column_letter(col_idx)].width = w

    # 人员配置数据
    # 格式: (角色, 职责范围, 对应模块, P1人月, P2人月, P3人月, 建议来源,
    #         初级日薪, 中级日薪, 高级日薪, 备注)
    STAFF = [
        (
            "全栈开发工程师（主力）",
            "Next.js 前端开发、PostgreSQL/Prisma 后端 API、权限系统、合同/财务/工单核心业务逻辑",
            "M0, M1, M2, M3, M5",
            6.0, 3.0, 2.0,
            "全职",
            800, 1500, 2500,
            "独立开发者方案核心角色；团队方案可拆为前端+后端各 1 人",
        ),
        (
            "前端开发工程师",
            "管理后台 UI 组件、CAD 可视化图层、NOI/WALE 看板图表、二房东填报端、响应式适配",
            "M1, M2, M3, M5",
            4.0, 2.0, 1.5,
            "全职",
            700, 1300, 2200,
            "团队方案需要；独立开发者方案由全栈兼任",
        ),
        (
            "后端开发工程师",
            "数据库设计、API 接口、账单引擎、递增规则计算、预警调度、数据导入导出",
            "M0, M2, M3, M5",
            4.0, 2.5, 1.5,
            "全职",
            700, 1400, 2500,
            "团队方案需要；独立开发者方案由全栈兼任",
        ),
        (
            "Flutter 移动端开发",
            "Flutter App（报修/工单/CAD 快查/推送通知）、租户自助门户 App",
            "M4, M6",
            2.0, 3.0, 0,
            "全职/外包",
            700, 1300, 2200,
            "Phase 1 工单模块 ~4 周；Phase 2 租户门户 ~7 周",
        ),
        (
            "微信小程序开发",
            "小程序精简版报修、扫码报修、在线出租展示页",
            "M4, M7",
            0.5, 1.5, 0,
            "兼职/外包",
            500, 1000, 1800,
            "Phase 1 仅精简报修(P2 可延期)；Phase 2 在线出租小程序",
        ),
        (
            "UI/UX 设计师",
            "设计系统搭建、管理后台页面设计、App 交互设计、二房东端表单、看板可视化",
            "M0, M1, M2, M3, M4, M5",
            1.5, 1.0, 0.5,
            "兼职/外包",
            500, 1100, 2000,
            "Phase 1 前 4 周集中设计，后续按需迭代",
        ),
        (
            "CAD/GIS 技术专家",
            ".dwg 转 SVG/PNG 技术调研与实现、图层叠加、矢量渲染优化",
            "M1",
            0.5, 0, 0,
            "外包/顾问",
            1000, 1800, 3000,
            "技术攻关期 ~2 周；可用开源方案(LibreCAD/OpenCTM)降低依赖",
        ),
        (
            "测试工程师",
            "功能测试、集成测试、UAT 协调、回归测试、性能基准测试",
            "M9, 全模块",
            1.5, 1.0, 0.5,
            "兼职/外包",
            400, 800, 1300,
            "Phase 1 最后 2 周集中测试+各模块完工后持续回归",
        ),
        (
            "DevOps / 运维",
            "CI/CD 流水线、服务器部署、数据库备份策略、监控告警、SSL/域名",
            "M0, M9",
            0.5, 0.5, 0,
            "兼职",
            700, 1300, 2200,
            "初期搭建 ~1 周，后续兼职维护；可由全栈兼任",
        ),
        (
            "产品经理 / 项目经理",
            "需求确认、优先级裁决、进度跟踪、UAT 验收协调、干系人沟通",
            "全模块",
            1.5, 1.0, 0.5,
            "兼职",
            700, 1500, 2500,
            "可由业务方兼任；独立开发者方案中开发者自行承担",
        ),
        (
            "数据录入 / 初始化",
            "639 套单元数据导入、500+ 在租合同录入、运营支出基线数据、二房东子租赁数据采集",
            "M9",
            1.0, 0, 0,
            "兼职/临时",
            250, 400, 400,
            "Phase 1 末期 3~4 周，与 UAT 并行；可由运营团队兼任",
        ),
        (
            "第三方集成工程师",
            "支付渠道对接（微信/支付宝）、电子签章（法大大/e签宝）、门禁系统（海康/华为云鼎）",
            "M6, M8",
            0, 2.0, 0,
            "外包",
            800, 1500, 2500,
            "Phase 2 专项；需第三方 SDK 对接经验",
        ),
    ]

    SOURCE_FILLS = {
        "全职": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "全职/外包": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "兼职/外包": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "兼职": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "兼职/临时": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "外包": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "外包/顾问": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    }

    LEVEL_FILLS = {
        "初级": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "中级": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "高级": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    }

    # ── 二线城市薪资 (成都/武汉/长沙/西安/郑州，约一线 65%-72%) ──
    TIER2_SALARY = {
        "全栈开发工程师（主力）": (550, 1000, 1700),
        "前端开发工程师":        (480, 900, 1500),
        "后端开发工程师":        (480, 950, 1700),
        "Flutter 移动端开发":    (480, 900, 1500),
        "微信小程序开发":        (350, 680, 1200),
        "UI/UX 设计师":         (350, 750, 1400),
        "CAD/GIS 技术专家":     (700, 1200, 2100),
        "测试工程师":            (280, 550, 900),
        "DevOps / 运维":        (480, 900, 1500),
        "产品经理 / 项目经理":    (480, 1000, 1700),
        "数据录入 / 初始化":      (180, 280, 280),
        "第三方集成工程师":       (550, 1000, 1700),
    }

    total_p1 = total_p2 = total_p3 = 0.0
    total_cost_j = total_cost_m = total_cost_s = 0.0
    for r_idx, (role, scope, modules, p1m, p2m, p3m, src, d_j, d_m, d_s, note) in enumerate(STAFF, 2):
        seq = r_idx - 1
        total_months = p1m + p2m + p3m
        m_j, m_m, m_s = d_j * 22, d_m * 22, d_s * 22
        cost_j = round(total_months * m_j)
        cost_m = round(total_months * m_m)
        cost_s = round(total_months * m_s)
        total_p1 += p1m
        total_p2 += p2m
        total_p3 += p3m
        total_cost_j += cost_j
        total_cost_m += cost_m
        total_cost_s += cost_s
        row_data = [seq, role, scope, modules, p1m, p2m, p3m, total_months, src,
                    d_j, m_j, d_m, m_m, d_s, m_s, cost_j, cost_m, cost_s, note]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CELL_ALIGN if c_idx in (2, 3, 4, 19) else CENTER_ALIGN
            if c_idx == 9 and src in SOURCE_FILLS:
                cell.fill = SOURCE_FILLS[src]
            if c_idx in (10, 11):
                cell.fill = LEVEL_FILLS["初级"]
            if c_idx in (12, 13):
                cell.fill = LEVEL_FILLS["中级"]
            if c_idx in (14, 15):
                cell.fill = LEVEL_FILLS["高级"]
            if c_idx == 16:
                cell.fill = LEVEL_FILLS["初级"]
            if c_idx == 17:
                cell.fill = LEVEL_FILLS["中级"]
            if c_idx == 18:
                cell.fill = LEVEL_FILLS["高级"]
            if c_idx in (10, 11, 12, 13, 14, 15, 16, 17, 18) and isinstance(val, (int, float)):
                cell.number_format = "#,##0"

    # 合计行
    total_all = total_p1 + total_p2 + total_p3
    tr = len(STAFF) + 2
    ws4.cell(row=tr, column=1, value="").border = THIN_BORDER
    ws4.cell(row=tr, column=2, value="合计").border = THIN_BORDER
    ws4.cell(row=tr, column=2).font = Font(bold=True)
    ws4.cell(row=tr, column=2).alignment = CENTER_ALIGN
    for c in range(3, 5):
        ws4.cell(row=tr, column=c, value="").border = THIN_BORDER
    for c_idx, val in [(5, total_p1), (6, total_p2), (7, total_p3), (8, total_all)]:
        cell = ws4.cell(row=tr, column=c_idx, value=val)
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        cell.font = Font(bold=True)
    for c in range(9, 16):
        ws4.cell(row=tr, column=c, value="").border = THIN_BORDER
    for c_idx, val, lvl in [(16, total_cost_j, "初级"), (17, total_cost_m, "中级"), (18, total_cost_s, "高级")]:
        cell = ws4.cell(row=tr, column=c_idx, value=round(val))
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        cell.font = Font(bold=True)
        cell.number_format = "#,##0"
        cell.fill = LEVEL_FILLS[lvl]
    ws4.cell(row=tr, column=19, value="").border = THIN_BORDER

    # 成本口径说明
    note_r4 = tr + 1
    ws4.cell(row=note_r4, column=1,
             value='※ “人员成本”口径：按各角色实际市场日薪 × 人月 × 22天计算，每个角色日薪不同，故总额与前三个Sheet（统一日薪估算）有差异，以本表为准。').alignment = CELL_ALIGN
    ws4.cell(row=note_r4, column=1).font = Font(italic=True, color="666666")

    ws4.freeze_panes = "A2"
    ws4.auto_filter.ref = ws4.dimensions

    # ── Sheet 5: 薪资对比 ────────────────────────────────────────
    ws5 = wb.create_sheet("薪资对比")
    comp_headers = [
        ("对比维度", 20), ("全用初级", 14), ("全用中级", 14), ("全用高级", 14),
        ("中级 vs 初级", 14), ("高级 vs 中级", 14),
    ]
    for col_idx, (h, w) in enumerate(comp_headers, 1):
        cell = ws5.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws5.column_dimensions[get_column_letter(col_idx)].width = w

    comp_rows = [
        ("总人力成本(元)", round(total_cost_j), round(total_cost_m), round(total_cost_s)),
        ("总人月", total_all, total_all, total_all),
        ("人均月薪(元)", round(total_cost_j / total_all) if total_all else 0,
         round(total_cost_m / total_all) if total_all else 0,
         round(total_cost_s / total_all) if total_all else 0),
    ]
    for r_idx, (label, v_j, v_m, v_s) in enumerate(comp_rows, 2):
        ws5.cell(row=r_idx, column=1, value=label).border = THIN_BORDER
        ws5.cell(row=r_idx, column=1).font = Font(bold=True)
        ws5.cell(row=r_idx, column=1).alignment = CELL_ALIGN
        for c_idx, (val, lvl) in enumerate([(v_j, "初级"), (v_m, "中级"), (v_s, "高级")], 2):
            cell = ws5.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.fill = LEVEL_FILLS[lvl]
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"
        # 差异列
        if isinstance(v_j, (int, float)) and v_j > 0:
            diff_mj = f"+{round((v_m - v_j) / v_j * 100)}%"
        else:
            diff_mj = "-"
        if isinstance(v_m, (int, float)) and v_m > 0:
            diff_sm = f"+{round((v_s - v_m) / v_m * 100)}%"
        else:
            diff_sm = "-"
        ws5.cell(row=r_idx, column=5, value=diff_mj).border = THIN_BORDER
        ws5.cell(row=r_idx, column=5).alignment = CENTER_ALIGN
        ws5.cell(row=r_idx, column=6, value=diff_sm).border = THIN_BORDER
        ws5.cell(row=r_idx, column=6).alignment = CENTER_ALIGN

    # 按角色明细对比
    r = 6
    ws5.cell(row=r, column=1, value="按角色明细对比").font = Font(bold=True, size=12)
    r += 1
    detail_headers = ["角色", "人月", "初级总成本", "中级总成本", "高级总成本", "中级溢价", "高级溢价"]
    for c_idx, h in enumerate(detail_headers, 1):
        cell = ws5.cell(row=r, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws5.column_dimensions[get_column_letter(1)].width = 25

    r += 1
    for role, _, _, p1m, p2m, p3m, _, d_j, d_m, d_s, _ in STAFF:
        tm = p1m + p2m + p3m
        cj = round(tm * d_j * 22)
        cm = round(tm * d_m * 22)
        cs = round(tm * d_s * 22)
        diff_mj = f"+¥{cm - cj:,}" if cj > 0 else "-"
        diff_sm = f"+¥{cs - cm:,}" if cm > 0 else "-"
        row_data = [role, tm, cj, cm, cs, diff_mj, diff_sm]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws5.cell(row=r, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if c_idx >= 2 else CELL_ALIGN
            if c_idx == 3:
                cell.fill = LEVEL_FILLS["初级"]
            if c_idx == 4:
                cell.fill = LEVEL_FILLS["中级"]
            if c_idx == 5:
                cell.fill = LEVEL_FILLS["高级"]
            if c_idx in (3, 4, 5) and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
        r += 1

    # 明细合计行
    ws5.cell(row=r, column=1, value="合计").font = Font(bold=True)
    ws5.cell(row=r, column=1).border = THIN_BORDER
    ws5.cell(row=r, column=2, value=total_all).border = THIN_BORDER
    ws5.cell(row=r, column=2).alignment = CENTER_ALIGN
    ws5.cell(row=r, column=2).font = Font(bold=True)
    for c_idx, (val, lvl) in enumerate([(total_cost_j, "初级"), (total_cost_m, "中级"), (total_cost_s, "高级")], 3):
        cell = ws5.cell(row=r, column=c_idx, value=round(val))
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        cell.font = Font(bold=True)
        cell.number_format = "#,##0"
        cell.fill = LEVEL_FILLS[lvl]
    diff_total_mj = f"+¥{round(total_cost_m - total_cost_j):,}"
    diff_total_sm = f"+¥{round(total_cost_s - total_cost_m):,}"
    ws5.cell(row=r, column=6, value=diff_total_mj).border = THIN_BORDER
    ws5.cell(row=r, column=6).alignment = CENTER_ALIGN
    ws5.cell(row=r, column=6).font = Font(bold=True)
    ws5.cell(row=r, column=7, value=diff_total_sm).border = THIN_BORDER
    ws5.cell(row=r, column=7).alignment = CENTER_ALIGN
    ws5.cell(row=r, column=7).font = Font(bold=True)

    # ── 区域 3: 按阶段分期汇总 ──
    r += 2
    ws5.cell(row=r, column=1, value="按阶段分期汇总").font = Font(bold=True, size=12)
    r += 1
    phase_cost_headers = ["阶段", "人月", "初级总成本", "中级总成本", "高级总成本",
                          "中级溢价", "高级溢价"]
    for c_idx, h in enumerate(phase_cost_headers, 1):
        cell = ws5.cell(row=r, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    r += 1

    # 按 Phase 汇总各角色的人月和成本
    phase_labels = ["Phase 1", "Phase 2", "Phase 3"]
    phase_month_keys = [
        lambda s: s[3],  # p1m
        lambda s: s[4],  # p2m
        lambda s: s[5],  # p3m
    ]
    grand_j = grand_m = grand_s = grand_months = 0
    for pi, phase_label in enumerate(phase_labels):
        pm_total = 0
        pcj = pcm = pcs = 0
        for role, _, _, p1m, p2m, p3m, _, d_j, d_m, d_s, _ in STAFF:
            pm = [p1m, p2m, p3m][pi]
            pm_total += pm
            pcj += round(pm * d_j * 22)
            pcm += round(pm * d_m * 22)
            pcs += round(pm * d_s * 22)
        grand_months += pm_total
        grand_j += pcj; grand_m += pcm; grand_s += pcs
        diff_mj = f"+¥{pcm - pcj:,}" if pcj > 0 else "-"
        diff_sm = f"+¥{pcs - pcm:,}" if pcm > 0 else "-"
        row_data = [phase_label, pm_total, pcj, pcm, pcs, diff_mj, diff_sm]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws5.cell(row=r, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if c_idx >= 2 else CELL_ALIGN
            if c_idx == 1 and phase_label in PHASE_FILLS:
                cell.fill = PHASE_FILLS[phase_label]
            if c_idx == 3:
                cell.fill = LEVEL_FILLS["初级"]
            if c_idx == 4:
                cell.fill = LEVEL_FILLS["中级"]
            if c_idx == 5:
                cell.fill = LEVEL_FILLS["高级"]
            if c_idx in (2, 3, 4, 5) and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
        r += 1

    # 阶段合计行
    ws5.cell(row=r, column=1, value="合计").font = Font(bold=True)
    ws5.cell(row=r, column=1).border = THIN_BORDER
    ws5.cell(row=r, column=2, value=grand_months).border = THIN_BORDER
    ws5.cell(row=r, column=2).alignment = CENTER_ALIGN
    ws5.cell(row=r, column=2).font = Font(bold=True)
    ws5.cell(row=r, column=2).number_format = "#,##0"
    for c_idx, (val, lvl) in enumerate([(grand_j, "初级"), (grand_m, "中级"), (grand_s, "高级")], 3):
        cell = ws5.cell(row=r, column=c_idx, value=round(val))
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN
        cell.font = Font(bold=True)
        cell.number_format = "#,##0"
        cell.fill = LEVEL_FILLS[lvl]
    diff_g_mj = f"+¥{round(grand_m - grand_j):,}"
    diff_g_sm = f"+¥{round(grand_s - grand_m):,}"
    ws5.cell(row=r, column=6, value=diff_g_mj).border = THIN_BORDER
    ws5.cell(row=r, column=6).alignment = CENTER_ALIGN
    ws5.cell(row=r, column=6).font = Font(bold=True)
    ws5.cell(row=r, column=7, value=diff_g_sm).border = THIN_BORDER
    ws5.cell(row=r, column=7).alignment = CENTER_ALIGN
    ws5.cell(row=r, column=7).font = Font(bold=True)

    for c_idx, w in enumerate([25, 8, 14, 14, 14, 14, 14], 1):
        ws5.column_dimensions[get_column_letter(c_idx)].width = max(
            ws5.column_dimensions[get_column_letter(c_idx)].width or 0, w)

    # ── Sheet 6: 交付方案对比 ────────────────────────────────────
    ws6 = wb.create_sheet("交付方案对比")

    # 三种团队方案定义
    # 格式: { 方案名: { "desc", "headcount", "parallel", "risk",
    #          "roles": [(角色, 人月, 薪资档)] } }
    # Phase 1 总工期 212 人天 = 42.4 人周
    P1_DAYS = phase_stats["Phase 1"]["days"]
    P2_DAYS = phase_stats["Phase 2"]["days"]

    PLANS = [
        {
            "name": "方案 A：独立开发者",
            "headcount": "1 人",
            "desc": "1 名全栈工程师串行开发全部模块，设计/测试/运维自行承担，数据录入由运营团队兼任",
            "p1_weeks": 32,
            "p2_weeks": 24,
            "parallel": "纯串行，模块逐个开发",
            "risk": "单点故障（病假/离职即停摆）；测试覆盖不足；设计品质受限",
            "advantage": "成本最低；沟通零开销；决策快",
            "roles": [
                ("全栈开发工程师", 7.5, "全职"),
                ("UI/UX 设计师", 0.5, "外包"),
                ("数据录入", 1.0, "临时"),
            ],
        },
        {
            "name": "方案 B：精简团队",
            "headcount": "3~4 人",
            "desc": "前后端各 1 人 + Flutter 1 人 + 设计师兼职，前后端可并行开发，模块间有部分重叠",
            "p1_weeks": 16,
            "p2_weeks": 14,
            "parallel": "前后端并行；M1/M2 可与 M4(App) 并行",
            "risk": "需协调前后端接口；关键人离职影响 1/3 进度",
            "advantage": "成本适中；工期减半；保留灵活性",
            "roles": [
                ("前端开发工程师", 5.0, "全职"),
                ("后端开发工程师", 5.0, "全职"),
                ("Flutter 移动端开发", 2.0, "全职"),
                ("UI/UX 设计师", 1.5, "兼职"),
                ("测试工程师", 1.0, "兼职"),
                ("数据录入", 1.0, "临时"),
            ],
        },
        {
            "name": "方案 C：完整团队",
            "headcount": "6~8 人",
            "desc": "前端 2 + 后端 2 + Flutter 1 + 设计 1 + 测试 1 + PM 1，多模块全面并行，专职测试保证质量",
            "p1_weeks": 10,
            "p2_weeks": 10,
            "parallel": "M1/M2/M3/M5 四模块并行开发；专职测试持续回归",
            "risk": "管理复杂度高；沟通成本大；前期招聘/磨合需时间",
            "advantage": "交付最快；测试充分；人员可替代性好",
            "roles": [
                ("前端开发工程师", 4.0, "全职"),
                ("前端开发工程师(二)", 3.5, "全职"),
                ("后端开发工程师", 4.0, "全职"),
                ("后端开发工程师(二)", 3.5, "全职"),
                ("Flutter 移动端开发", 2.0, "全职"),
                ("UI/UX 设计师", 1.5, "全职"),
                ("测试工程师", 2.0, "全职"),
                ("产品经理/项目经理", 1.5, "兼职"),
                ("数据录入", 1.0, "临时"),
            ],
        },
    ]

    # 角色→薪资档映射 (从 STAFF 数据中提取)
    ROLE_SALARY = {}
    ROLE_SALARY_T2 = {}
    for role, _, _, _, _, _, _, d_j, d_m, d_s, _ in STAFF:
        ROLE_SALARY[role] = (d_j, d_m, d_s)
        ROLE_SALARY_T2[role] = TIER2_SALARY.get(role, (round(d_j * 0.68), round(d_m * 0.68), round(d_s * 0.68)))
    # 补充复合角色的映射
    ROLE_SALARY["前端开发工程师(二)"] = ROLE_SALARY.get("前端开发工程师", (700, 1300, 2200))
    ROLE_SALARY["后端开发工程师(二)"] = ROLE_SALARY.get("后端开发工程师", (700, 1400, 2500))
    ROLE_SALARY["全栈开发工程师"] = ROLE_SALARY.get("全栈开发工程师（主力）", (800, 1500, 2500))
    ROLE_SALARY["产品经理/项目经理"] = ROLE_SALARY.get("产品经理 / 项目经理", (700, 1500, 2500))
    ROLE_SALARY.setdefault("数据录入", ROLE_SALARY.get("数据录入 / 初始化", (250, 400, 400)))
    # 二线复合映射
    ROLE_SALARY_T2["前端开发工程师(二)"] = ROLE_SALARY_T2.get("前端开发工程师", (480, 900, 1500))
    ROLE_SALARY_T2["后端开发工程师(二)"] = ROLE_SALARY_T2.get("后端开发工程师", (480, 950, 1700))
    ROLE_SALARY_T2["全栈开发工程师"] = ROLE_SALARY_T2.get("全栈开发工程师（主力）", (550, 1000, 1700))
    ROLE_SALARY_T2["产品经理/项目经理"] = ROLE_SALARY_T2.get("产品经理 / 项目经理", (480, 1000, 1700))
    ROLE_SALARY_T2.setdefault("数据录入", ROLE_SALARY_T2.get("数据录入 / 初始化", (180, 280, 280)))

    PLAN_FILLS = {
        "方案 A：独立开发者": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "方案 B：精简团队": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "方案 C：完整团队": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    }

    # ── 区域 1: 方案总览对比表 ──
    overview_headers = [
        ("对比维度", 22), ("方案 A：独立开发者", 24),
        ("方案 B：精简团队", 24), ("方案 C：完整团队", 24),
    ]
    for c_idx, (h, w) in enumerate(overview_headers, 1):
        cell = ws6.cell(row=1, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws6.column_dimensions[get_column_letter(c_idx)].width = w

    overview_rows = [
        ("团队规模", [p["headcount"] for p in PLANS]),
        ("Phase 1 工期(周)", [p["p1_weeks"] for p in PLANS]),
        ("Phase 2 工期(周)", [p["p2_weeks"] for p in PLANS]),
        ("总交付周期(周)", [p["p1_weeks"] + p["p2_weeks"] for p in PLANS]),
        ("并行度", [p["parallel"] for p in PLANS]),
        ("优势", [p["advantage"] for p in PLANS]),
        ("风险", [p["risk"] for p in PLANS]),
        ("方案说明", [p["desc"] for p in PLANS]),
    ]

    for r_idx, (label, values) in enumerate(overview_rows, 2):
        ws6.cell(row=r_idx, column=1, value=label).border = THIN_BORDER
        ws6.cell(row=r_idx, column=1).font = Font(bold=True)
        ws6.cell(row=r_idx, column=1).alignment = CELL_ALIGN
        for c_idx, val in enumerate(values, 2):
            cell = ws6.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CELL_ALIGN
            cell.fill = PLAN_FILLS[PLANS[c_idx - 2]["name"]]

    # ── 区域 2: 成本矩阵 (3方案 × 3薪资档 × 2城市 = 18 种组合) ──
    r = len(overview_rows) + 3
    CITY_FILL_T1 = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    CITY_FILL_T2 = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws6.cell(row=r, column=1, value="成本矩阵（3 方案 × 3 薪资档 × 2 城市等级）").font = Font(bold=True, size=12)
    r += 1

    cost_matrix_headers = ["方案", "城市等级", "团队规模",
                           "初级总成本(元)", "中级总成本(元)", "高级总成本(元)",
                           "初级月均(元)", "中级月均(元)", "高级月均(元)"]
    for c_idx, h in enumerate(cost_matrix_headers, 1):
        cell = ws6.cell(row=r, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    r += 1

    for plan in PLANS:
        for city_label, sal_dict, city_fill in [
            ("一线", ROLE_SALARY, CITY_FILL_T1),
            ("二线", ROLE_SALARY_T2, CITY_FILL_T2),
        ]:
            cost_j = cost_m = cost_s = 0
            for role_name, months, _ in plan["roles"]:
                sal = sal_dict.get(role_name, (500, 1000, 1800))
                cost_j += round(months * sal[0] * 22)
                cost_m += round(months * sal[1] * 22)
                cost_s += round(months * sal[2] * 22)

            total_weeks = plan["p1_weeks"] + plan["p2_weeks"]
            total_months_cal = total_weeks / 4.33
            burn_j = round(cost_j / total_months_cal) if total_months_cal > 0 else 0
            burn_m = round(cost_m / total_months_cal) if total_months_cal > 0 else 0
            burn_s = round(cost_s / total_months_cal) if total_months_cal > 0 else 0

            row_data = [plan["name"], city_label, plan["headcount"],
                        cost_j, cost_m, cost_s, burn_j, burn_m, burn_s]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws6.cell(row=r, column=c_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN if c_idx >= 2 else CELL_ALIGN
                if c_idx == 1:
                    cell.fill = PLAN_FILLS[plan["name"]]
                if c_idx == 2:
                    cell.fill = city_fill
                if c_idx in (4, 7):
                    cell.fill = LEVEL_FILLS["初级"]
                if c_idx in (5, 8):
                    cell.fill = LEVEL_FILLS["中级"]
                if c_idx in (6, 9):
                    cell.fill = LEVEL_FILLS["高级"]
                if c_idx >= 4 and isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
            r += 1

    # ── 区域 3: 各方案角色明细（一线 / 二线对比） ──
    for plan in PLANS:
        r += 1
        ws6.cell(row=r, column=1, value=plan["name"]).font = Font(bold=True, size=11)
        ws6.cell(row=r, column=1).fill = PLAN_FILLS[plan["name"]]
        r += 1

        role_headers = ["角色", "人月", "来源",
                        "一线初级(元)", "一线中级(元)", "一线高级(元)",
                        "二线初级(元)", "二线中级(元)", "二线高级(元)"]
        for c_idx, h in enumerate(role_headers, 1):
            cell = ws6.cell(row=r, column=c_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        r += 1

        plan_t1_j = plan_t1_m = plan_t1_s = 0
        plan_t2_j = plan_t2_m = plan_t2_s = 0
        for role_name, months, src in plan["roles"]:
            sal1 = ROLE_SALARY.get(role_name, (500, 1000, 1800))
            sal2 = ROLE_SALARY_T2.get(role_name, (350, 680, 1200))
            c1j = round(months * sal1[0] * 22)
            c1m = round(months * sal1[1] * 22)
            c1s = round(months * sal1[2] * 22)
            c2j = round(months * sal2[0] * 22)
            c2m = round(months * sal2[1] * 22)
            c2s = round(months * sal2[2] * 22)
            plan_t1_j += c1j; plan_t1_m += c1m; plan_t1_s += c1s
            plan_t2_j += c2j; plan_t2_m += c2m; plan_t2_s += c2s
            row_data = [role_name, months, src, c1j, c1m, c1s, c2j, c2m, c2s]
            for c_idx, val in enumerate(row_data, 1):
                cell = ws6.cell(row=r, column=c_idx, value=val)
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN if c_idx >= 2 else CELL_ALIGN
                if c_idx in (4, 7):
                    cell.fill = LEVEL_FILLS["初级"]
                if c_idx in (5, 8):
                    cell.fill = LEVEL_FILLS["中级"]
                if c_idx in (6, 9):
                    cell.fill = LEVEL_FILLS["高级"]
                if c_idx >= 4 and isinstance(val, (int, float)):
                    cell.number_format = "#,##0"
            r += 1

        # 方案小计
        ws6.cell(row=r, column=1, value="小计").font = Font(bold=True)
        ws6.cell(row=r, column=1).border = THIN_BORDER
        tm = sum(m for _, m, _ in plan["roles"])
        ws6.cell(row=r, column=2, value=tm).border = THIN_BORDER
        ws6.cell(row=r, column=2).alignment = CENTER_ALIGN
        ws6.cell(row=r, column=2).font = Font(bold=True)
        ws6.cell(row=r, column=3, value="").border = THIN_BORDER
        for c_idx, val in [(4, plan_t1_j), (5, plan_t1_m), (6, plan_t1_s),
                           (7, plan_t2_j), (8, plan_t2_m), (9, plan_t2_s)]:
            cell = ws6.cell(row=r, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.font = Font(bold=True)
            cell.number_format = "#,##0"
            if c_idx in (4, 7):
                cell.fill = LEVEL_FILLS["初级"]
            if c_idx in (5, 8):
                cell.fill = LEVEL_FILLS["中级"]
            if c_idx in (6, 9):
                cell.fill = LEVEL_FILLS["高级"]
        r += 1

    for c_idx, w in enumerate([24, 8, 10, 10, 14, 10, 14, 10, 14], 1):
        cur = ws6.column_dimensions[get_column_letter(c_idx)].width or 0
        ws6.column_dimensions[get_column_letter(c_idx)].width = max(cur, w)

    ws6.freeze_panes = "A2"

    # ── Sheet 7: 城市薪资对比 ────────────────────────────────────
    ws7 = wb.create_sheet("城市薪资对比")

    # ── 区域 1: 各角色一线/二线日薪对比 ──
    city_headers = [
        ("角色", 25), ("来源", 10),
        ("一线初级日薪", 12), ("二线初级日薪", 12), ("差异", 8),
        ("一线中级日薪", 12), ("二线中级日薪", 12), ("差异", 8),
        ("一线高级日薪", 12), ("二线高级日薪", 12), ("差异", 8),
    ]
    for c_idx, (h, w) in enumerate(city_headers, 1):
        cell = ws7.cell(row=1, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws7.column_dimensions[get_column_letter(c_idx)].width = w

    for r_idx, (role, scope, modules, p1m, p2m, p3m, src, d_j, d_m, d_s, note) in enumerate(STAFF, 2):
        t2 = TIER2_SALARY.get(role, (round(d_j * 0.68), round(d_m * 0.68), round(d_s * 0.68)))
        diff_j = f"-{round((1 - t2[0] / d_j) * 100)}%" if d_j > 0 else "-"
        diff_m = f"-{round((1 - t2[1] / d_m) * 100)}%" if d_m > 0 else "-"
        diff_s = f"-{round((1 - t2[2] / d_s) * 100)}%" if d_s > 0 else "-"
        row_data = [role, src, d_j, t2[0], diff_j, d_m, t2[1], diff_m, d_s, t2[2], diff_s]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws7.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if c_idx >= 2 else CELL_ALIGN
            if c_idx in (3, 6, 9):
                cell.fill = CITY_FILL_T1
            if c_idx in (4, 7, 10):
                cell.fill = CITY_FILL_T2
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"

    # ── 区域 2: 一线/二线月薪参考 ──
    r2 = len(STAFF) + 3
    ws7.cell(row=r2, column=1, value="月薪参考（日薪 × 22 天）").font = Font(bold=True, size=12)
    r2 += 1

    month_headers = [
        "角色",
        "一线初级月薪", "二线初级月薪",
        "一线中级月薪", "二线中级月薪",
        "一线高级月薪", "二线高级月薪",
    ]
    for c_idx, h in enumerate(month_headers, 1):
        cell = ws7.cell(row=r2, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    r2 += 1

    for role, _, _, _, _, _, _, d_j, d_m, d_s, _ in STAFF:
        t2 = TIER2_SALARY.get(role, (round(d_j * 0.68), round(d_m * 0.68), round(d_s * 0.68)))
        row_data = [role,
                    d_j * 22, t2[0] * 22,
                    d_m * 22, t2[1] * 22,
                    d_s * 22, t2[2] * 22]
        for c_idx, val in enumerate(row_data, 1):
            cell = ws7.cell(row=r2, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if c_idx >= 2 else CELL_ALIGN
            if c_idx in (2, 4, 6):
                cell.fill = CITY_FILL_T1
            if c_idx in (3, 5, 7):
                cell.fill = CITY_FILL_T2
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"
        r2 += 1

    # ── 区域 3: 全部人员总成本对比 (一线 vs 二线 × 初/中/高) ──
    r3 = r2 + 1
    ws7.cell(row=r3, column=1, value="全员总成本对比（一线 vs 二线）").font = Font(bold=True, size=12)
    r3 += 1

    sum_headers = ["维度", "全用初级", "全用中级", "全用高级"]
    for c_idx, h in enumerate(sum_headers, 1):
        cell = ws7.cell(row=r3, column=c_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    r3 += 1

    # 一线总成本 (已算好 total_cost_j/m/s)
    t2_cost_j = t2_cost_m = t2_cost_s = 0
    for role, _, _, p1m, p2m, p3m, _, d_j, d_m, d_s, _ in STAFF:
        t2 = TIER2_SALARY.get(role, (round(d_j * 0.68), round(d_m * 0.68), round(d_s * 0.68)))
        tm = p1m + p2m + p3m
        t2_cost_j += round(tm * t2[0] * 22)
        t2_cost_m += round(tm * t2[1] * 22)
        t2_cost_s += round(tm * t2[2] * 22)

    cost_compare = [
        ("一线城市（北上广深杭）", round(total_cost_j), round(total_cost_m), round(total_cost_s)),
        ("二线城市（成都/武汉/长沙/西安）", t2_cost_j, t2_cost_m, t2_cost_s),
        ("节省金额(元)", round(total_cost_j) - t2_cost_j,
         round(total_cost_m) - t2_cost_m, round(total_cost_s) - t2_cost_s),
    ]
    # 节省比例
    save_pct = []
    for i in range(3):
        t1 = [round(total_cost_j), round(total_cost_m), round(total_cost_s)][i]
        t2v = [t2_cost_j, t2_cost_m, t2_cost_s][i]
        save_pct.append(f"-{round((t1 - t2v) / t1 * 100)}%" if t1 > 0 else "-")
    cost_compare.append(("节省比例", save_pct[0], save_pct[1], save_pct[2]))

    for label, v1, v2, v3 in cost_compare:
        ws7.cell(row=r3, column=1, value=label).border = THIN_BORDER
        ws7.cell(row=r3, column=1).font = Font(bold=True)
        ws7.cell(row=r3, column=1).alignment = CELL_ALIGN
        for c_idx, (val, lvl) in enumerate([(v1, "初级"), (v2, "中级"), (v3, "高级")], 2):
            cell = ws7.cell(row=r3, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.fill = LEVEL_FILLS[lvl]
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"
        r3 += 1

    # ── 区域 4: 薪资数据说明 ──
    r4 = r3 + 1
    ws7.cell(row=r4, column=1, value="数据说明").font = Font(bold=True, size=12)
    r4 += 1
    notes = [
        "1. 一线城市指北京、上海、广州、深圳、杭州，二线指成都、武汉、长沙、西安、郑州等",
        "2. 日薪为外包/合同工含税报价（含社保/管理费），非纯到手工资",
        "3. 月薪 = 日薪 × 22 天（每月平均工作日），年薪 ≈ 月薪 × 12",
        "4. 数据来源：BOSS 直聘/拉勾/猎聘 2025-2026 外包岗位行情 + 行业访谈",
        "5. 实际薪资受具体城市、公司规模、项目复杂度影响，表中为外包市场中位数",
        "6. 全职雇佣需额外加 35%~45% 的五险一金企业部分（表中未包含）",
    ]
    for note in notes:
        ws7.cell(row=r4, column=1, value=note).alignment = CELL_ALIGN
        r4 += 1

    ws7.freeze_panes = "A2"

    wb.save(OUT_FILE)
    print(f"✅ 导出完成: {OUT_FILE}")
    print(f"   共 {len(wb.sheetnames)} 个 Sheet: {', '.join(wb.sheetnames)}")
    print(f"   需求总数: {len(REQUIREMENTS)} 条")
    print(f"   人员角色: {len(STAFF)} 个")
    for phase, stats in phase_stats.items():
        print(f"   {phase}: {stats['count']} 条, {stats['days']} 人天, 一线中级¥{round(stats['days'] * DAY_COST_T1):,} / 一线高级¥{round(stats['days'] * DAY_COST_T1_HIGH):,}")


if __name__ == "__main__":
    main()
